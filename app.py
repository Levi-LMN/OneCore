from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import datetime, date, timedelta, timezone
import json
from decimal import Decimal
import re

# Import models from your models.py file
from models import db, User, Category, Size, ExpenseCategory, Product, ProductVariant, Expense, DailyStock, Sale, \
    DailySummary, AuditLog, StockPurchase

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here-change-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///liquor_store.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
}
# Make "today" available globally in all templates
app.jinja_env.globals['today'] = date.today()

db.init_app(app)





# Helper Functions
def get_current_user():
    """Get current user"""
    if 'user_id' in session:
        return db.session.get(User, session['user_id'])
    return None


def create_audit_log(action, table_name, record_id=None, old_values=None, new_values=None, changes_summary=None):
    """Create an audit log entry"""
    try:
        current_user = get_current_user()
        if not current_user:
            return

        ip_address = request.environ.get('HTTP_X_FORWARDED_FOR',
                                         request.environ.get('REMOTE_ADDR', ''))
        user_agent = request.environ.get('HTTP_USER_AGENT', '')

        old_json = json.dumps(old_values, default=str) if old_values else None
        new_json = json.dumps(new_values, default=str) if new_values else None

        audit_log = AuditLog(
            user_id=current_user.id,
            action=action,
            table_name=table_name,
            record_id=record_id,
            old_values=old_json,
            new_values=new_json,
            changes_summary=changes_summary,
            ip_address=ip_address[:45] if ip_address else None,
            user_agent=user_agent[:500] if user_agent else None
        )

        db.session.add(audit_log)

    except Exception as e:
        app.logger.error(f"Error creating audit log: {str(e)}")


def get_changes_summary(old_values, new_values):
    """Generate a human-readable summary of changes"""
    if not old_values or not new_values:
        return "Record created or deleted"

    changes = []
    friendly_names = {
        'opening_stock': 'Opening Stock',
        'additions': 'Stock Additions',
        'closing_stock': 'Closing Stock',
        'current_stock': 'Current Stock',
        'min_stock_level': 'Minimum Stock Level',
        'base_buying_price': 'Base Buying Price',
        'selling_price': 'Selling Price',
        'conversion_factor': 'Conversion Factor',
        'base_unit': 'Base Unit',
        'total_sales': 'Total Sales',
        'total_expenses': 'Total Expenses',
        'cash_amount': 'Cash Amount',
        'mpesa_amount': 'M-Pesa Amount',
        'credit_amount': 'Credit Amount',
        'quantity': 'Quantity',
        'unit_price': 'Unit Price',
        'total_amount': 'Total Amount',
        'amount': 'Amount',
        'description': 'Description',
        'category_id': 'Category',
        'size_id': 'Size',
        'product_id': 'Product',
        'variant_id': 'Product Variant',
        'expense_category_id': 'Category',
        'is_active': 'Status',
        'full_name': 'Full Name',
        'username': 'Username',
        'email': 'Email',
        'role': 'Role',
        'name': 'Name',
        'sort_order': 'Sort Order'
    }

    for key, new_value in new_values.items():
        if key in old_values:
            old_value = old_values[key]
            if str(old_value) != str(new_value):
                field_name = friendly_names.get(key, key.replace('_', ' ').title())

                if key in ['opening_stock', 'additions', 'closing_stock', 'current_stock', 'quantity', 'min_stock_level', 'conversion_factor']:
                    changes.append(f"{field_name}: {old_value} → {new_value}")
                elif 'amount' in key.lower() or 'price' in key.lower():
                    changes.append(f"{field_name}: KES {old_value} → KES {new_value}")
                elif key == 'is_active':
                    status_old = "Active" if old_value else "Inactive"
                    status_new = "Active" if new_value else "Inactive"
                    changes.append(f"{field_name}: {status_old} → {status_new}")
                else:
                    changes.append(f"{field_name}: '{old_value}' → '{new_value}'")

    return "; ".join(changes) if changes else "No changes detected"


def safe_float(value, default=0.0):
    """Safely convert a value to float"""
    try:
        return float(value) if value not in [None, ''] else default
    except (ValueError, TypeError):
        return default


def validate_email(email):
    """Validate email format"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email))




def update_daily_stock_sales(product_id, stock_date):
    """Update daily stock sales from actual sales (now calculating base units from variants)"""
    # Calculate total base units sold from all variant sales
    total_base_units_sold = db.session.query(
        db.func.coalesce(db.func.sum(Sale.quantity * ProductVariant.conversion_factor), 0)
    ).join(ProductVariant, Sale.variant_id == ProductVariant.id).filter(
        ProductVariant.product_id == product_id,
        Sale.sale_date == stock_date
    ).scalar()

    # Update daily stock
    daily_stock = get_or_create_daily_stock(product_id, stock_date)
    daily_stock.sales_quantity = total_base_units_sold
    daily_stock.calculate_closing_stock()

    return daily_stock


def update_daily_summary(target_date):
    """Update or create daily summary for a specific date"""
    try:
        # Calculate sales for the date
        sales_data = db.session.query(
            db.func.count(Sale.id).label('transaction_count'),
            db.func.coalesce(db.func.sum(Sale.total_amount), 0).label('total_sales'),
            db.func.coalesce(db.func.sum(
                Sale.total_amount - (Product.base_buying_price * ProductVariant.conversion_factor * Sale.quantity)),
                             0).label('total_profit')
        ).join(ProductVariant).join(Product).filter(
            Sale.sale_date == target_date
        ).first()

        # Calculate expenses for the date
        expenses_data = db.session.query(
            db.func.count(Expense.id).label('expense_count'),
            db.func.coalesce(db.func.sum(Expense.amount), 0).label('total_expenses')
        ).filter(
            Expense.expense_date == target_date
        ).first()

        # Update or create daily summary
        summary = DailySummary.query.filter_by(date=target_date).first()
        if not summary:
            summary = DailySummary(date=target_date)
            db.session.add(summary)

        summary.total_transactions = sales_data.transaction_count or 0
        summary.total_sales = sales_data.total_sales or 0
        summary.total_profit = sales_data.total_profit or 0
        summary.total_expenses = expenses_data.total_expenses or 0
        summary.expense_count = expenses_data.expense_count or 0
        summary.net_profit = (sales_data.total_profit or 0) - (expenses_data.total_expenses or 0)
        summary.updated_at = datetime.now(timezone.utc)

        return summary

    except Exception as e:
        app.logger.error(f"Error updating daily summary for {target_date}: {str(e)}")
        return None


# Authentication decorators
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login'))

        user = db.session.get(User, session['user_id'])
        if not user or user.role not in ['admin', 'manager']:
            flash('Access denied. Admin privileges required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function


# AUTHENTICATION ROUTES
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email'].strip().lower()
        password = request.form['password']
        remember = request.form.get('remember') == 'on'

        user = User.query.filter_by(email=email, is_active=True).first()

        if user and user.check_password(password):
            session.update({
                'user_id': user.id,
                'username': user.username,
                'user_role': user.role,
                'user_full_name': user.full_name,
                'user_email': user.email
            })

            # Set session to be permanent if "Remember Me" is checked
            if remember:
                session.permanent = True
                # Set session lifetime to 30 days (configurable)
                app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)
            else:
                session.permanent = False

            user.last_login = datetime.now(timezone.utc)

            create_audit_log(
                action='LOGIN',
                table_name='user',
                record_id=user.id,
                changes_summary=f"User {user.full_name} logged in successfully" +
                               (" (Remember Me enabled)" if remember else "")
            )

            db.session.commit()
            flash(f'Welcome back, {user.full_name}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid email or password', 'error')

    return render_template('login.html')

@app.route('/logout')
def logout():
    current_user = get_current_user()
    if current_user:
        create_audit_log(
            action='LOGOUT',
            table_name='user',
            record_id=current_user.id,
            changes_summary=f"User {current_user.full_name} logged out"
        )
        db.session.commit()

    session.clear()
    flash('You have been logged out successfully', 'info')
    return redirect(url_for('login'))


# DASHBOARD
@app.route('/')
@app.route('/dashboard')
@login_required
def dashboard():
    current_user = get_current_user()

    # Get selected date from query params, default to today
    selected_date_str = request.args.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
    except ValueError:
        selected_date = date.today()

    today = date.today()

    # Base queries with role-based filtering
    sales_base_query = db.session.query(Sale).join(ProductVariant).join(Product)
    expenses_base_query = db.session.query(Expense)

    # Filter for attendants - only their own records
    if current_user.role not in ['admin', 'manager']:
        sales_base_query = sales_base_query.filter(Sale.attendant_id == current_user.id)
        expenses_base_query = expenses_base_query.filter(Expense.recorded_by == current_user.id)

    # TODAY'S STATISTICS
    today_sales_query = sales_base_query.filter(Sale.sale_date == selected_date)
    today_expenses_query = expenses_base_query.filter(Expense.expense_date == selected_date)

    # Calculate today's metrics
    today_sales_data = today_sales_query.with_entities(
        db.func.count(Sale.id).label('transaction_count'),
        db.func.coalesce(db.func.sum(Sale.total_amount), 0).label('total_sales'),
        db.func.coalesce(db.func.sum(Sale.total_amount - (Product.base_buying_price * ProductVariant.conversion_factor * Sale.quantity)), 0).label('gross_profit')
        if current_user.role in ['admin', 'manager'] else db.func.coalesce(db.func.sum(Sale.total_amount), 0).label('gross_profit')
    ).first()

    today_expenses_data = today_expenses_query.with_entities(
        db.func.count(Expense.id).label('expense_count'),
        db.func.coalesce(db.func.sum(Expense.amount), 0).label('total_expenses')
    ).first()

    # Today's stats object
    today_stats = {
        'total_transactions': today_sales_data.transaction_count or 0,
        'total_sales': today_sales_data.total_sales or 0,
        'total_expenses': today_expenses_data.total_expenses or 0,
        'expense_count': today_expenses_data.expense_count or 0,
    }

    # Add profit calculations for admin/manager
    if current_user.role in ['admin', 'manager']:
        today_stats['gross_profit'] = today_sales_data.gross_profit or 0
        today_stats['net_profit'] = today_stats['gross_profit'] - today_stats['total_expenses']
    else:
        today_stats['gross_profit'] = 0
        today_stats['net_profit'] = 0

    # MONTHLY STATISTICS (for admin/manager)
    month_stats = {'total_sales': 0, 'total_profit': 0, 'total_expenses': 0, 'net_profit': 0}
    if current_user.role in ['admin', 'manager']:
        month_start = selected_date.replace(day=1)

        month_sales_data = sales_base_query.filter(
            Sale.sale_date >= month_start,
            Sale.sale_date <= selected_date
        ).with_entities(
            db.func.coalesce(db.func.sum(Sale.total_amount), 0).label('total_sales'),
            db.func.coalesce(db.func.sum(Sale.total_amount - (Product.base_buying_price * ProductVariant.conversion_factor * Sale.quantity)), 0).label('gross_profit')
        ).first()

        month_expenses_total = expenses_base_query.filter(
            Expense.expense_date >= month_start,
            Expense.expense_date <= selected_date
        ).with_entities(
            db.func.coalesce(db.func.sum(Expense.amount), 0)
        ).scalar() or 0

        month_stats = {
            'total_sales': month_sales_data.total_sales or 0,
            'total_profit': month_sales_data.gross_profit or 0,
            'total_expenses': month_expenses_total,
            'net_profit': (month_sales_data.gross_profit or 0) - month_expenses_total
        }

    # STOCK ALERTS
    low_stock_products = Product.query.filter(
        Product.current_stock <= Product.min_stock_level
    ).order_by(Product.current_stock.asc()).limit(10).all()

    low_stock_count = Product.query.filter(
        Product.current_stock <= Product.min_stock_level
    ).count()

    # RECENT SALES
    recent_sales_query = today_sales_query.order_by(Sale.timestamp.desc()).limit(5)
    recent_sales = recent_sales_query.all()

    # TOP SELLING PRODUCTS (for selected date) - Updated for variant system
    top_products_base = db.session.query(
        Product.name.label('product_name'),
        Category.name.label('category_name'),
        db.func.sum(Sale.quantity).label('total_quantity'),
        db.func.sum(Sale.total_amount).label('total_sales')
    ).select_from(Sale).join(ProductVariant).join(Product).join(Category, Product.category_id == Category.id).filter(
        Sale.sale_date == selected_date
    )

    # Add profit calculation for admin/manager
    if current_user.role in ['admin', 'manager']:
        top_products_query = top_products_base.add_columns(
            db.func.sum(Sale.total_amount - (Product.base_buying_price * ProductVariant.conversion_factor * Sale.quantity)).label('total_profit')
        )
    else:
        top_products_query = top_products_base.add_columns(
            db.func.sum(Sale.total_amount * 0).label('total_profit')  # Always 0 for attendants
        )

    # Filter for attendants
    if current_user.role not in ['admin', 'manager']:
        top_products_query = top_products_query.filter(Sale.attendant_id == current_user.id)

    top_products = top_products_query.group_by(
        Product.id, Product.name, Category.name
    ).order_by(
        db.text('total_sales DESC')
    ).limit(5).all()

    return render_template('dashboard.html',
                           current_user=current_user,
                           today=today,
                           selected_date=selected_date,
                           today_stats=today_stats,
                           month_stats=month_stats,
                           low_stock_count=low_stock_count,
                           stock_alerts=low_stock_products,
                           recent_sales=recent_sales,
                           top_products=top_products)


# USER MANAGEMENT ROUTES
@app.route('/add_user', methods=['GET', 'POST'])
@admin_required
def add_user():
    if request.method == 'POST':
        username = request.form['username'].strip()
        email = request.form['email'].strip().lower()
        full_name = request.form['full_name'].strip()
        password = request.form['password']
        role = request.form['role']

        if not all([username, email, full_name, password, role]):
            flash('All fields are required!', 'error')
            return redirect(url_for('add_user'))

        existing_user = User.query.filter(
            (User.username == username) | (User.email == email)
        ).first()

        if existing_user:
            if existing_user.username == username:
                flash('Username already exists!', 'error')
            else:
                flash('Email already exists!', 'error')
            return redirect(url_for('add_user'))

        if not validate_email(email):
            flash('Please enter a valid email address!', 'error')
            return redirect(url_for('add_user'))

        user = User(
            username=username,
            email=email,
            full_name=full_name,
            role=role
        )
        user.set_password(password)

        db.session.add(user)
        db.session.flush()

        new_values = user.to_dict()
        new_values.pop('password_hash', None)

        create_audit_log(
            action='CREATE',
            table_name='user',
            record_id=user.id,
            new_values=new_values,
            changes_summary=f"New user created: {full_name} ({username}) with role {role}"
        )

        db.session.commit()
        flash(f'User {full_name} created successfully!', 'success')
        return redirect(url_for('users'))

    return render_template('admin/add_user.html')


@app.route('/users')
@admin_required
def users():
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template('admin/users.html', users=users)


@app.route('/toggle_user_status/<int:user_id>', methods=['POST'])
@admin_required
def toggle_user_status(user_id):
    user = db.session.get(User, user_id)
    if not user:
        flash('User not found!', 'error')
        return redirect(url_for('users'))

    old_values = user.to_dict()
    user.is_active = not user.is_active
    new_values = user.to_dict()

    status = "activated" if user.is_active else "deactivated"

    create_audit_log(
        action='UPDATE',
        table_name='user',
        record_id=user.id,
        old_values=old_values,
        new_values=new_values,
        changes_summary=f"User {user.full_name} has been {status}"
    )

    db.session.commit()
    flash(f'User {user.full_name} has been {status}', 'success')
    return redirect(url_for('users'))


@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
@admin_required
def edit_user(user_id):
    user = db.session.get(User, user_id)
    if not user:
        flash('User not found!', 'error')
        return redirect(url_for('users'))

    if request.method == 'POST':
        old_values = user.to_dict()

        username = request.form['username'].strip()
        email = request.form['email'].strip().lower()
        full_name = request.form['full_name'].strip()
        role = request.form['role']

        # Check if username exists (excluding current user)
        existing = User.query.filter(User.username == username, User.id != user_id).first()
        if existing:
            flash('Username already exists!', 'error')
            return render_template('admin/edit_user.html', user=user)

        # Check if email exists (excluding current user)
        existing = User.query.filter(User.email == email, User.id != user_id).first()
        if existing:
            flash('Email already exists!', 'error')
            return render_template('admin/edit_user.html', user=user)

        if not validate_email(email):
            flash('Please enter a valid email address!', 'error')
            return render_template('admin/edit_user.html', user=user)

        user.username = username
        user.email = email
        user.full_name = full_name
        user.role = role

        new_values = user.to_dict()
        changes_summary = get_changes_summary(old_values, new_values)

        create_audit_log(
            action='UPDATE',
            table_name='user',
            record_id=user.id,
            old_values=old_values,
            new_values=new_values,
            changes_summary=f"User updated: {user.full_name} - {changes_summary}"
        )

        db.session.commit()
        flash(f'User "{user.full_name}" updated successfully!', 'success')
        return redirect(url_for('users'))

    return render_template('admin/edit_user.html', user=user)


@app.route('/admin/reset_user_password/<int:user_id>', methods=['POST'])
@admin_required
def reset_user_password(user_id):
    user = db.session.get(User, user_id)
    if not user:
        flash('User not found!', 'error')
        return redirect(url_for('users'))

    new_password = request.form['new_password']
    confirm_password = request.form['confirm_password']
    notify_user = request.form.get('notify_user') == 'on'

    if new_password != confirm_password:
        flash('Passwords do not match!', 'error')
        return redirect(url_for('users'))

    if len(new_password) < 6:
        flash('Password must be at least 6 characters long!', 'error')
        return redirect(url_for('users'))

    user.set_password(new_password)

    create_audit_log(
        action='UPDATE',
        table_name='user',
        record_id=user.id,
        changes_summary=f"Password reset for user {user.full_name} by admin"
    )

    db.session.commit()

    flash_msg = f'Password for "{user.full_name}" has been reset successfully!'
    if notify_user:
        flash_msg += ' User should be notified of the change.'

    flash(flash_msg, 'success')
    return redirect(url_for('users'))


@app.route('/delete_user/<int:user_id>', methods=['POST'])
@admin_required
def delete_user(user_id):
    current_user = get_current_user()

    # Prevent self-deletion
    if user_id == current_user.id:
        flash('You cannot delete your own account!', 'error')
        return redirect(url_for('users'))

    user = db.session.get(User, user_id)
    if not user:
        flash('User not found!', 'error')
        return redirect(url_for('users'))

    user_name = user.full_name

    # 🔍 Check all models that reference User
    has_sales = Sale.query.filter_by(attendant_id=user_id).first()
    has_expenses = Expense.query.filter_by(recorded_by=user_id).first()
    has_purchases = StockPurchase.query.filter_by(recorded_by=user_id).first()
    has_categories = Category.query.filter_by(created_by=user_id).first()
    has_products = Product.query.filter_by(created_by=user_id).first()
    has_variants = ProductVariant.query.filter_by(created_by=user_id).first()
    has_sizes = Size.query.filter_by(created_by=user_id).first()
    has_expense_categories = ExpenseCategory.query.filter_by(created_by=user_id).first()
    has_daily_stocks = DailyStock.query.filter_by(updated_by=user_id).first()
    has_summaries = DailySummary.query.filter_by(last_updated_by=user_id).first()
    has_audits = AuditLog.query.filter_by(user_id=user_id).first()

    if any([
        has_sales, has_expenses, has_purchases,
        has_categories, has_products, has_variants,
        has_sizes, has_expense_categories,
        has_daily_stocks, has_summaries, has_audits
    ]):
        flash(
            f'User "{user_name}" cannot be deleted because they have linked records '
            f'(sales, expenses, stock updates, etc.). Please deactivate them instead.',
            'error'
        )
        return redirect(url_for('users'))

    # ✅ Safe to delete (no related records)
    old_values = user.to_dict()

    # Commit audit log first
    create_audit_log(
        action='DELETE',
        table_name='user',
        record_id=user_id,
        old_values=old_values,
        changes_summary=f"User deleted: {user_name}"
    )
    db.session.commit()

    # Delete user
    db.session.delete(user)
    db.session.commit()

    flash(f'User "{user_name}" deleted successfully!', 'success')
    return redirect(url_for('users'))


# CATEGORY MANAGEMENT ROUTES
@app.route('/categories')
@admin_required
def categories():
    categories = Category.query.order_by(Category.name).all()
    return render_template('categories/categories.html', categories=categories)


@app.route('/add_category', methods=['GET', 'POST'])
@admin_required
def add_category():
    if request.method == 'POST':
        name = request.form['name'].strip()
        description = request.form.get('description', '').strip()
        current_user = get_current_user()

        if not name:
            flash('Category name is required!', 'error')
            return redirect(url_for('add_category'))

        existing_category = Category.query.filter_by(name=name).first()
        if existing_category:
            flash('Category already exists!', 'error')
            return redirect(url_for('add_category'))

        category = Category(
            name=name,
            description=description if description else None,
            created_by=current_user.id
        )

        db.session.add(category)
        db.session.flush()

        create_audit_log(
            action='CREATE',
            table_name='category',
            record_id=category.id,
            new_values=category.to_dict(),
            changes_summary=f"New category created: {name}"
        )

        db.session.commit()
        flash(f'Category "{name}" created successfully!', 'success')
        return redirect(url_for('categories'))

    return render_template('categories/add_category.html')


@app.route('/edit_category/<int:category_id>', methods=['GET', 'POST'])
@admin_required
def edit_category(category_id):
    category = db.session.get(Category, category_id)
    if not category:
        flash('Category not found!', 'error')
        return redirect(url_for('categories'))

    if request.method == 'POST':
        old_values = category.to_dict()

        name = request.form['name'].strip()
        description = request.form.get('description', '').strip()

        # Check if name already exists (excluding current category)
        existing = Category.query.filter(Category.name == name, Category.id != category_id).first()
        if existing:
            flash('Category name already exists!', 'error')
            return render_template('categories/edit_category.html', category=category)

        category.name = name
        category.description = description if description else None

        new_values = category.to_dict()
        changes_summary = get_changes_summary(old_values, new_values)

        create_audit_log(
            action='UPDATE',
            table_name='category',
            record_id=category.id,
            old_values=old_values,
            new_values=new_values,
            changes_summary=f"Category updated: {category.name} - {changes_summary}"
        )

        db.session.commit()
        flash(f'Category "{category.name}" updated successfully!', 'success')
        return redirect(url_for('categories'))

    return render_template('categories/edit_category.html', category=category)


@app.route('/toggle_category_status/<int:category_id>', methods=['POST'])
@admin_required
def toggle_category_status(category_id):
    category = db.session.get(Category, category_id)
    if not category:
        flash('Category not found!', 'error')
        return redirect(url_for('categories'))

    old_values = category.to_dict()
    category.is_active = not category.is_active
    new_values = category.to_dict()

    status = "activated" if category.is_active else "deactivated"

    create_audit_log(
        action='UPDATE',
        table_name='category',
        record_id=category.id,
        old_values=old_values,
        new_values=new_values,
        changes_summary=f"Category {category.name} has been {status}"
    )

    db.session.commit()
    flash(f'Category "{category.name}" has been {status}', 'success')
    return redirect(url_for('categories'))


@app.route('/delete_category/<int:category_id>', methods=['POST'])
@admin_required
def delete_category(category_id):
    category = db.session.get(Category, category_id)
    if not category:
        flash('Category not found!', 'error')
        return redirect(url_for('categories'))

    # Check if category has products
    if category.products:
        flash(f'Cannot delete category "{category.name}" - it has {len(category.products)} product(s) associated with it!', 'error')
        return redirect(url_for('categories'))

    old_values = category.to_dict()
    category_name = category.name

    create_audit_log(
        action='DELETE',
        table_name='category',
        record_id=category_id,
        old_values=old_values,
        changes_summary=f"Category deleted: {category_name}"
    )

    db.session.delete(category)
    db.session.commit()
    flash(f'Category "{category_name}" deleted successfully!', 'success')
    return redirect(url_for('categories'))

# SIZE MANAGEMENT ROUTES
@app.route('/sizes')
@admin_required
def sizes():
    sizes = Size.query.order_by(Size.sort_order, Size.name).all()
    return render_template('sizes/sizes.html', sizes=sizes)


@app.route('/add_size', methods=['GET', 'POST'])
@admin_required
def add_size():
    if request.method == 'POST':
        name = request.form['name'].strip()
        description = request.form.get('description', '').strip()
        sort_order = safe_float(request.form.get('sort_order', 0))
        current_user = get_current_user()

        if not name:
            flash('Size name is required!', 'error')
            return redirect(url_for('add_size'))

        existing_size = Size.query.filter_by(name=name).first()
        if existing_size:
            flash('Size already exists!', 'error')
            return redirect(url_for('add_size'))

        size = Size(
            name=name,
            description=description if description else None,
            sort_order=int(sort_order),
            created_by=current_user.id
        )

        db.session.add(size)
        db.session.flush()

        create_audit_log(
            action='CREATE',
            table_name='size',
            record_id=size.id,
            new_values=size.to_dict(),
            changes_summary=f"New size created: {name}"
        )

        db.session.commit()
        flash(f'Size "{name}" created successfully!', 'success')
        return redirect(url_for('sizes'))

    return render_template('sizes/add_size.html')


@app.route('/edit_size/<int:size_id>', methods=['GET', 'POST'])
@admin_required
def edit_size(size_id):
    size = db.session.get(Size, size_id)
    if not size:
        flash('Size not found!', 'error')
        return redirect(url_for('sizes'))

    if request.method == 'POST':
        old_values = size.to_dict()

        name = request.form['name'].strip()
        description = request.form.get('description', '').strip()
        sort_order = safe_float(request.form.get('sort_order', 0))

        # Check if name already exists (excluding current size)
        existing = Size.query.filter(Size.name == name, Size.id != size_id).first()
        if existing:
            flash('Size name already exists!', 'error')
            return render_template('sizes/edit_size.html', size=size)

        size.name = name
        size.description = description if description else None
        size.sort_order = int(sort_order)

        new_values = size.to_dict()
        changes_summary = get_changes_summary(old_values, new_values)

        create_audit_log(
            action='UPDATE',
            table_name='size',
            record_id=size.id,
            old_values=old_values,
            new_values=new_values,
            changes_summary=f"Size updated: {size.name} - {changes_summary}"
        )

        db.session.commit()
        flash(f'Size "{size.name}" updated successfully!', 'success')
        return redirect(url_for('sizes'))

    return render_template('sizes/edit_size.html', size=size)


@app.route('/toggle_size_status/<int:size_id>', methods=['POST'])
@admin_required
def toggle_size_status(size_id):
    size = db.session.get(Size, size_id)
    if not size:
        flash('Size not found!', 'error')
        return redirect(url_for('sizes'))

    old_values = size.to_dict()
    size.is_active = not size.is_active
    new_values = size.to_dict()

    status = "activated" if size.is_active else "deactivated"

    create_audit_log(
        action='UPDATE',
        table_name='size',
        record_id=size.id,
        old_values=old_values,
        new_values=new_values,
        changes_summary=f"Size {size.name} has been {status}"
    )

    db.session.commit()
    flash(f'Size "{size.name}" has been {status}', 'success')
    return redirect(url_for('sizes'))


@app.route('/delete_size/<int:size_id>', methods=['POST'])
@admin_required
def delete_size(size_id):
    size = db.session.get(Size, size_id)
    if not size:
        flash('Size not found!', 'error')
        return redirect(url_for('sizes'))

    # Check if size has product variants
    if size.product_variants:
        flash(f'Cannot delete size "{size.name}" - it has {len(size.product_variants)} variant(s) associated with it!',
              'error')
        return redirect(url_for('sizes'))

    old_values = size.to_dict()
    size_name = size.name

    create_audit_log(
        action='DELETE',
        table_name='size',
        record_id=size_id,
        old_values=old_values,
        changes_summary=f"Size deleted: {size_name}"
    )

    db.session.delete(size)
    db.session.commit()
    flash(f'Size "{size_name}" deleted successfully!', 'success')
    return redirect(url_for('sizes'))


# EXPENSE CATEGORY ROUTES
@app.route('/expense_categories')
@admin_required
def expense_categories():
    expense_categories = ExpenseCategory.query.order_by(ExpenseCategory.name).all()
    return render_template('expenses/expense_categories.html', expense_categories=expense_categories)


@app.route('/add_expense_category', methods=['GET', 'POST'])
@admin_required
def add_expense_category():
    if request.method == 'POST':
        name = request.form['name'].strip()
        description = request.form.get('description', '').strip()
        current_user = get_current_user()

        if not name:
            flash('Expense category name is required!', 'error')
            return redirect(url_for('add_expense_category'))

        existing_category = ExpenseCategory.query.filter_by(name=name).first()
        if existing_category:
            flash('Expense category already exists!', 'error')
            return redirect(url_for('add_expense_category'))

        expense_category = ExpenseCategory(
            name=name,
            description=description if description else None,
            created_by=current_user.id
        )

        db.session.add(expense_category)
        db.session.flush()

        create_audit_log(
            action='CREATE',
            table_name='expense_category',
            record_id=expense_category.id,
            new_values=expense_category.to_dict(),
            changes_summary=f"New expense category created: {name}"
        )

        db.session.commit()
        flash(f'Expense category "{name}" created successfully!', 'success')
        return redirect(url_for('expense_categories'))

    return render_template('expenses/add_expense_category.html')


@app.route('/edit_expense_category/<int:category_id>', methods=['GET', 'POST'])
@admin_required
def edit_expense_category(category_id):
    category = db.session.get(ExpenseCategory, category_id)
    if not category:
        flash('Expense category not found!', 'error')
        return redirect(url_for('expense_categories'))

    if request.method == 'POST':
        old_values = category.to_dict()

        name = request.form['name'].strip()
        description = request.form.get('description', '').strip()

        # Check if name already exists (excluding current category)
        existing = ExpenseCategory.query.filter(ExpenseCategory.name == name, ExpenseCategory.id != category_id).first()
        if existing:
            flash('Expense category name already exists!', 'error')
            return render_template('expenses/edit_expense_category.html', category=category)

        category.name = name
        category.description = description if description else None

        new_values = category.to_dict()
        changes_summary = get_changes_summary(old_values, new_values)

        create_audit_log(
            action='UPDATE',
            table_name='expense_category',
            record_id=category.id,
            old_values=old_values,
            new_values=new_values,
            changes_summary=f"Expense category updated: {category.name} - {changes_summary}"
        )

        db.session.commit()
        flash(f'Expense category "{category.name}" updated successfully!', 'success')
        return redirect(url_for('expense_categories'))

    return render_template('expenses/edit_expense_category.html', category=category)


@app.route('/toggle_expense_category_status/<int:category_id>', methods=['POST'])
@admin_required
def toggle_expense_category_status(category_id):
    category = db.session.get(ExpenseCategory, category_id)
    if not category:
        flash('Expense category not found!', 'error')
        return redirect(url_for('expense_categories'))

    old_values = category.to_dict()
    category.is_active = not category.is_active
    new_values = category.to_dict()

    status = "activated" if category.is_active else "deactivated"

    create_audit_log(
        action='UPDATE',
        table_name='expense_category',
        record_id=category.id,
        old_values=old_values,
        new_values=new_values,
        changes_summary=f"Expense category {category.name} has been {status}"
    )

    db.session.commit()
    flash(f'Expense category "{category.name}" has been {status}', 'success')
    return redirect(url_for('expense_categories'))


@app.route('/delete_expense_category/<int:category_id>', methods=['POST'])
@admin_required
def delete_expense_category(category_id):
    category = db.session.get(ExpenseCategory, category_id)
    if not category:
        flash('Expense category not found!', 'error')
        return redirect(url_for('expense_categories'))

    # Check if category has expenses
    if category.expenses:
        flash(
            f'Cannot delete expense category "{category.name}" - it has {len(category.expenses)} expense(s) associated with it!',
            'error')
        return redirect(url_for('expense_categories'))

    old_values = category.to_dict()
    category_name = category.name

    create_audit_log(
        action='DELETE',
        table_name='expense_category',
        record_id=category_id,
        old_values=old_values,
        changes_summary=f"Expense category deleted: {category_name}"
    )

    db.session.delete(category)
    db.session.commit()
    flash(f'Expense category "{category_name}" deleted successfully!', 'success')
    return redirect(url_for('expense_categories'))


# PRODUCT MANAGEMENT ROUTES
@app.route('/add_product', methods=['GET', 'POST'])
@admin_required
def add_product():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        category_id = request.form.get('category_id', type=int)
        base_unit = request.form.get('base_unit', 'bottle').strip()
        base_buying_price = safe_float(request.form.get('base_buying_price'))
        opening_stock = safe_float(request.form.get('opening_stock'))
        min_stock_level = safe_float(request.form.get('min_stock_level', 5))

        current_user = get_current_user()

        if not all([name, category_id, base_unit]) or base_buying_price <= 0:
            flash('Please fill in all required fields with valid values!', 'error')
            return redirect(url_for('add_product'))

        if Product.query.filter_by(name=name).first():
            flash('Product already exists!', 'error')
            return redirect(url_for('add_product'))

        # Validate category exists and is active
        category = db.session.get(Category, category_id)
        if not category or not category.is_active:
            flash('Invalid category selected!', 'error')
            return redirect(url_for('add_product'))

        product = Product(
            name=name,
            category_id=category_id,
            base_unit=base_unit,
            base_buying_price=base_buying_price,
            current_stock=opening_stock,
            min_stock_level=min_stock_level,
            created_by=current_user.id,
            last_stock_update=datetime.now(timezone.utc)
        )

        db.session.add(product)
        db.session.flush()

        create_audit_log(
            action='CREATE',
            table_name='product',
            record_id=product.id,
            new_values=product.to_dict(),
            changes_summary=f"New product created: {name} (Category: {category.name}, Base Unit: {base_unit}, Base Price: KES {base_buying_price}, Initial Stock: {opening_stock})"
        )

        if opening_stock > 0:
            today = date.today()
            initial_stock = DailyStock(
                product_id=product.id,
                date=today,
                opening_stock=opening_stock,
                additions=0,
                sales_quantity=0,
                closing_stock=opening_stock,
                updated_by=current_user.id
            )
            db.session.add(initial_stock)

        db.session.commit()

        success_msg = f'Product added successfully'
        if opening_stock > 0:
            success_msg += f' with opening stock of {opening_stock} {base_unit}s!'
        else:
            success_msg += '! Remember to add product variants and set initial stock.'

        flash(success_msg, 'success' if opening_stock > 0 else 'info')
        return redirect(url_for('products'))

    categories = Category.query.filter_by(is_active=True).order_by(Category.name).all()
    return render_template('products/add_product.html', categories=categories)


@app.route('/products')
@login_required
def products():
    category_id = request.args.get('category_id', 'all')
    stock_status = request.args.get('stock_status', 'all')

    query = db.session.query(Product, Category).select_from(Product) \
        .join(Category, Product.category_id == Category.id)

    if category_id != 'all':
        query = query.filter(Product.category_id == int(category_id))

    if stock_status == 'low_stock':
        query = query.filter(Product.current_stock <= Product.min_stock_level, Product.current_stock > 0)
    elif stock_status == 'out_of_stock':
        query = query.filter(Product.current_stock <= 0)
    elif stock_status == 'good_stock':
        query = query.filter(Product.current_stock > Product.min_stock_level)

    products_data = query.order_by(Category.name, Product.name).all()

    # Get variants for each product
    products_with_variants = []
    for product, category in products_data:
        variants = product.get_active_variants()
        products_with_variants.append((product, category, variants))

    categories = Category.query.filter_by(is_active=True).order_by(Category.name).all()

    return render_template('products/products.html',
                           products_data=products_with_variants,
                           categories=categories,
                           selected_category_id=category_id,
                           selected_stock_status=stock_status)


@app.route('/edit_product/<int:product_id>', methods=['GET', 'POST'])
@admin_required
def edit_product(product_id):
    product = db.session.get(Product, product_id)
    if not product:
        flash('Product not found!', 'error')
        return redirect(url_for('products'))

    if request.method == 'POST':
        old_values = product.to_dict()

        name = request.form['name'].strip()
        category_id = int(request.form['category_id'])
        base_unit = request.form['base_unit'].strip()
        base_buying_price = safe_float(request.form['base_buying_price'])
        min_stock_level = safe_float(request.form['min_stock_level'])

        # Check if name already exists (excluding current product)
        existing = Product.query.filter(Product.name == name, Product.id != product_id).first()
        if existing:
            flash('Product name already exists!', 'error')
            categories = Category.query.filter_by(is_active=True).order_by(Category.name).all()
            return render_template('products/edit_product.html', product=product, categories=categories)

        # Validate category
        category = db.session.get(Category, category_id)
        if not category or not category.is_active:
            flash('Invalid category selected!', 'error')
            categories = Category.query.filter_by(is_active=True).order_by(Category.name).all()
            return render_template('products/edit_product.html', product=product, categories=categories)

        product.name = name
        product.category_id = category_id
        product.base_unit = base_unit
        product.base_buying_price = base_buying_price
        product.min_stock_level = min_stock_level

        new_values = product.to_dict()
        changes_summary = get_changes_summary(old_values, new_values)

        create_audit_log(
            action='UPDATE',
            table_name='product',
            record_id=product.id,
            old_values=old_values,
            new_values=new_values,
            changes_summary=f"Product updated: {product.name} - {changes_summary}"
        )

        db.session.commit()
        flash(f'Product "{product.name}" updated successfully!', 'success')
        return redirect(url_for('products'))

    categories = Category.query.filter_by(is_active=True).order_by(Category.name).all()
    return render_template('products/edit_product.html', product=product, categories=categories)


@app.route('/delete_product/<int:product_id>', methods=['POST'])
@admin_required
def delete_product(product_id):
    product = db.session.get(Product, product_id)
    if not product:
        flash('Product not found!', 'error')
        return redirect(url_for('products'))

    # Check if product has sales
    has_sales = Sale.query.join(ProductVariant).filter(ProductVariant.product_id == product_id).first()
    if has_sales:
        flash(f'Cannot delete product "{product.name}" - it has sales records associated with it!', 'error')
        return redirect(url_for('products'))

    old_values = product.to_dict()
    product_name = product.name

    # Delete all variants first
    for variant in product.variants:
        db.session.delete(variant)

    create_audit_log(
        action='DELETE',
        table_name='product',
        record_id=product_id,
        old_values=old_values,
        changes_summary=f"Product deleted: {product_name}"
    )

    db.session.delete(product)
    db.session.commit()
    flash(f'Product "{product_name}" and all its variants deleted successfully!', 'success')
    return redirect(url_for('products'))


# PRODUCT VARIANT ROUTES
@app.route('/add_variant/<int:product_id>', methods=['GET', 'POST'])
@admin_required
def add_variant(product_id):
    product = db.session.get(Product, product_id)
    if not product:
        flash('Product not found!', 'error')
        return redirect(url_for('products'))

    if request.method == 'POST':
        size_id = request.form.get('size_id', type=int)
        selling_price = safe_float(request.form.get('selling_price'))
        conversion_factor = safe_float(request.form.get('conversion_factor', 1.0))

        current_user = get_current_user()

        if not all([size_id]) or selling_price <= 0 or conversion_factor <= 0:
            flash('Please fill in all required fields with valid values!', 'error')
            return redirect(url_for('add_variant', product_id=product_id))

        # Validate size exists and is active
        size = db.session.get(Size, size_id)
        if not size or not size.is_active:
            flash('Invalid size selected!', 'error')
            return redirect(url_for('add_variant', product_id=product_id))

        # Check if variant already exists
        existing_variant = ProductVariant.query.filter_by(
            product_id=product_id,
            size_id=size_id
        ).first()
        if existing_variant:
            flash(f'Variant for {size.name} already exists!', 'error')
            return redirect(url_for('add_variant', product_id=product_id))

        variant = ProductVariant(
            product_id=product_id,
            size_id=size_id,
            selling_price=selling_price,
            conversion_factor=conversion_factor,
            created_by=current_user.id
        )

        db.session.add(variant)
        db.session.flush()

        create_audit_log(
            action='CREATE',
            table_name='product_variant',
            record_id=variant.id,
            new_values=variant.to_dict(),
            changes_summary=f"New variant created: {product.name} - {size.name} (Price: KES {selling_price}, Factor: {conversion_factor})"
        )

        db.session.commit()
        flash(f'Product variant "{variant.get_display_name()}" created successfully!', 'success')
        return redirect(url_for('product_variants', product_id=product_id))

    sizes = Size.query.filter_by(is_active=True).order_by(Size.sort_order, Size.name).all()
    return render_template('products/add_variant.html', product=product, sizes=sizes)


@app.route('/product_variants/<int:product_id>')
@login_required
def product_variants(product_id):
    product = db.session.get(Product, product_id)
    if not product:
        flash('Product not found!', 'error')
        return redirect(url_for('products'))

    variants = db.session.query(ProductVariant, Size).join(Size).filter(
        ProductVariant.product_id == product_id
    ).order_by(Size.sort_order).all()

    return render_template('products/product_variants.html', product=product, variants=variants)


@app.route('/edit_variant/<int:variant_id>', methods=['GET', 'POST'])
@admin_required
def edit_variant(variant_id):
    variant = db.session.get(ProductVariant, variant_id)
    if not variant:
        flash('Product variant not found!', 'error')
        return redirect(url_for('products'))

    if request.method == 'POST':
        old_values = variant.to_dict()

        selling_price = safe_float(request.form.get('selling_price'))
        conversion_factor = safe_float(request.form.get('conversion_factor', 1.0))

        if selling_price <= 0 or conversion_factor <= 0:
            flash('Please provide valid selling price and conversion factor!', 'error')
            sizes = Size.query.filter_by(is_active=True).order_by(Size.sort_order, Size.name).all()
            return render_template('products/edit_variant.html', variant=variant, sizes=sizes)

        variant.selling_price = selling_price
        variant.conversion_factor = conversion_factor

        new_values = variant.to_dict()
        changes_summary = get_changes_summary(old_values, new_values)

        create_audit_log(
            action='UPDATE',
            table_name='product_variant',
            record_id=variant.id,
            old_values=old_values,
            new_values=new_values,
            changes_summary=f"Variant updated: {variant.get_display_name()} - {changes_summary}"
        )

        db.session.commit()
        flash(f'Product variant "{variant.get_display_name()}" updated successfully!', 'success')
        return redirect(url_for('product_variants', product_id=variant.product_id))

    sizes = Size.query.filter_by(is_active=True).order_by(Size.sort_order, Size.name).all()
    return render_template('products/edit_variant.html', variant=variant, sizes=sizes)


@app.route('/toggle_variant_status/<int:variant_id>', methods=['POST'])
@admin_required
def toggle_variant_status(variant_id):
    variant = db.session.get(ProductVariant, variant_id)
    if not variant:
        flash('Product variant not found!', 'error')
        return redirect(url_for('products'))

    old_values = variant.to_dict()
    variant.is_active = not variant.is_active
    new_values = variant.to_dict()

    status = "activated" if variant.is_active else "deactivated"

    create_audit_log(
        action='UPDATE',
        table_name='product_variant',
        record_id=variant.id,
        old_values=old_values,
        new_values=new_values,
        changes_summary=f"Variant {variant.get_display_name()} has been {status}"
    )

    db.session.commit()
    flash(f'Variant "{variant.get_display_name()}" has been {status}', 'success')
    return redirect(url_for('product_variants', product_id=variant.product_id))


@app.route('/delete_variant/<int:variant_id>', methods=['POST'])
@admin_required
def delete_variant(variant_id):
    variant = db.session.get(ProductVariant, variant_id)
    if not variant:
        flash('Product variant not found!', 'error')
        return redirect(url_for('products'))

    # Check if variant has sales
    has_sales = Sale.query.filter_by(variant_id=variant_id).first()
    if has_sales:
        flash(f'Cannot delete variant "{variant.get_display_name()}" - it has sales records associated with it!', 'error')
        return redirect(url_for('product_variants', product_id=variant.product_id))

    old_values = variant.to_dict()
    variant_name = variant.get_display_name()
    product_id = variant.product_id

    create_audit_log(
        action='DELETE',
        table_name='product_variant',
        record_id=variant_id,
        old_values=old_values,
        changes_summary=f"Variant deleted: {variant_name}"
    )

    db.session.delete(variant)
    db.session.commit()
    flash(f'Variant "{variant_name}" deleted successfully!', 'success')
    return redirect(url_for('product_variants', product_id=product_id))


# SALES ROUTES
@app.route('/sales')
@login_required
def sales():
    selected_date_str = request.args.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
    except ValueError:
        selected_date = date.today()

    # Check if a specific date was requested (not today)
    stick_to_date = selected_date_str != date.today().strftime('%Y-%m-%d')

    current_user = get_current_user()

    # Updated query for variant system
    sales_query = db.session.query(Sale, ProductVariant, Product, User, Category, Size).select_from(Sale) \
        .join(ProductVariant, Sale.variant_id == ProductVariant.id) \
        .join(Product, ProductVariant.product_id == Product.id) \
        .join(User, Sale.attendant_id == User.id) \
        .join(Category, Product.category_id == Category.id) \
        .join(Size, ProductVariant.size_id == Size.id) \
        .filter(Sale.sale_date == selected_date)

    if current_user.role not in ['admin', 'manager']:
        sales_query = sales_query.filter(Sale.attendant_id == current_user.id)

    sales_data = sales_query.order_by(Sale.timestamp.desc()).all()

    # Calculate totals
    sale_amounts = [(sale.original_amount, sale.discount_amount, sale.total_amount, sale.get_profit())
                    for sale, _, _, _, _, _ in sales_data]

    totals = {
        'original': sum(amounts[0] for amounts in sale_amounts),
        'discount': sum(amounts[1] for amounts in sale_amounts),
        'sales': sum(amounts[2] for amounts in sale_amounts),
        'profit': sum(amounts[3] for amounts in sale_amounts)
    }

    # Get available variants
    variants = db.session.query(ProductVariant, Product, Category, Size).select_from(ProductVariant) \
        .join(Product, ProductVariant.product_id == Product.id) \
        .join(Category, Product.category_id == Category.id) \
        .join(Size, ProductVariant.size_id == Size.id) \
        .filter(
        ProductVariant.is_active == True,
        Product.current_stock > 0,
        Category.is_active == True,
        Size.is_active == True
    ).order_by(Category.name, Product.name, Size.sort_order).all()

    today = date.today()

    return render_template(
        'sales/sales.html',
        sales_data=sales_data,
        selected_date=selected_date,
        stick_to_date=stick_to_date,
        total_original=totals['original'],
        total_discount=totals['discount'],
        total_sales=totals['sales'],
        total_profit=totals['profit'],
        variants=variants,
        current_user=current_user,
        today=today
    )


@app.route('/add_sale', methods=['POST'])
@login_required
def add_sale():
    try:
        current_user = get_current_user()

        # Check if we should return to a specific date
        return_date = request.form.get('return_date')

        # Extract form data
        variant_id = int(request.form.get('variant_id'))
        quantity = safe_float(request.form.get('quantity', 0))
        unit_price = safe_float(request.form.get('unit_price', 0))

        # Discount fields
        discount_type = request.form.get('discount_type', 'none')
        discount_value = safe_float(request.form.get('discount_value', 0))
        discount_reason = request.form.get('discount_reason', '').strip()

        # Payment fields
        cash_amount = safe_float(request.form.get('cash_amount', 0))
        mpesa_amount = safe_float(request.form.get('mpesa_amount', 0))
        credit_amount = safe_float(request.form.get('credit_amount', 0))
        customer_name = request.form.get('customer_name', '').strip()
        notes = request.form.get('notes', '').strip()
        sale_date_str = request.form.get('sale_date')

        # Validate required fields
        if quantity <= 0 or unit_price <= 0:
            flash('Please fill in all required fields correctly.', 'error')
            redirect_date = return_date if return_date else date.today().strftime('%Y-%m-%d')
            return redirect(url_for('sales', date=redirect_date))

        # Parse sale date
        try:
            sale_date = datetime.strptime(sale_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            flash('Invalid sale date format.', 'error')
            redirect_date = return_date if return_date else date.today().strftime('%Y-%m-%d')
            return redirect(url_for('sales', date=redirect_date))

        # Get variant
        variant = db.session.get(ProductVariant, variant_id)
        if not variant or not variant.is_active:
            flash('Selected product variant not found or inactive.', 'error')
            redirect_date = return_date if return_date else date.today().strftime('%Y-%m-%d')
            return redirect(url_for('sales', date=redirect_date))

        # Check if we can sell the requested quantity
        if not variant.can_sell_quantity(quantity):
            available = variant.get_available_stock_in_variant_units()
            flash(f'Insufficient stock! Only {available} units of {variant.get_display_name()} available.', 'error')
            redirect_date = return_date if return_date else date.today().strftime('%Y-%m-%d')
            return redirect(url_for('sales', date=redirect_date))

        # Validate discount permissions
        if discount_type != 'none' and discount_value > 0:
            max_discount_map = {'attendant': 10, 'manager': 25, 'admin': 100}
            max_discount = max_discount_map.get(current_user.role, 0)

            if discount_type == 'percentage' and discount_value > max_discount:
                flash(f'You can only give up to {max_discount}% discount. Contact admin for higher discounts.', 'error')
                redirect_date = return_date if return_date else date.today().strftime('%Y-%m-%d')
                return redirect(url_for('sales', date=redirect_date))

            if not discount_reason:
                flash('Please provide a reason for the discount.', 'error')
                redirect_date = return_date if return_date else date.today().strftime('%Y-%m-%d')
                return redirect(url_for('sales', date=redirect_date))

        # Calculate amounts
        original_amount = quantity * unit_price
        sale = Sale(
            variant_id=variant.id,
            quantity=quantity,
            unit_price=unit_price,
            original_amount=original_amount,
            discount_type=discount_type,
            discount_value=discount_value,
            discount_reason=discount_reason if discount_reason else None,
            sale_date=sale_date,
            attendant_id=current_user.id,
            cash_amount=cash_amount,
            mpesa_amount=mpesa_amount,
            credit_amount=credit_amount,
            customer_name=customer_name if customer_name else None,
            notes=notes if notes else None
        )

        sale.calculate_discount()

        # Validate payment
        payment_total = cash_amount + mpesa_amount + credit_amount
        if payment_total < sale.total_amount:
            flash(f'Insufficient payment! Total: KES {sale.total_amount:,.2f}, Paid: KES {payment_total:,.2f}', 'error')
            redirect_date = return_date if return_date else date.today().strftime('%Y-%m-%d')
            return redirect(url_for('sales', date=redirect_date))

        if credit_amount > 0 and not customer_name:
            flash('Customer name is required for credit sales.', 'error')
            redirect_date = return_date if return_date else date.today().strftime('%Y-%m-%d')
            return redirect(url_for('sales', date=redirect_date))

        # Handle excess payment
        if payment_total > sale.total_amount:
            change_amount = payment_total - sale.total_amount
            excess_note = f"Change given: KES {change_amount:.2f}"
            sale.notes = f"{sale.notes}. {excess_note}" if sale.notes else excess_note

        # Determine payment method
        payment_methods = []
        if cash_amount > 0: payment_methods.append('cash')
        if mpesa_amount > 0: payment_methods.append('mpesa')
        if credit_amount > 0: payment_methods.append('credit')

        sale.payment_method = 'mixed' if len(payment_methods) > 1 else (
            payment_methods[0] if payment_methods else 'cash')

        # Reduce stock in base units
        base_units_needed = quantity * variant.conversion_factor
        if not variant.product.reduce_stock(base_units_needed):
            flash(f'Failed to reduce stock for {variant.get_display_name()}. Please try again.', 'error')
            redirect_date = return_date if return_date else date.today().strftime('%Y-%m-%d')
            return redirect(url_for('sales', date=redirect_date))

        db.session.add(sale)
        db.session.flush()

        # Create audit log
        changes_summary = f"Sale: {variant.get_display_name()} x{quantity} @ KES {unit_price}"
        if sale.discount_amount > 0:
            changes_summary += f" (Discount: -KES {sale.discount_amount:.2f}, Final: KES {sale.total_amount:.2f})"
            if sale.discount_reason:
                changes_summary += f" - Reason: {sale.discount_reason}"
        else:
            changes_summary += f" = KES {sale.total_amount:.2f}"

        if customer_name:
            changes_summary += f" (Customer: {customer_name})"

        create_audit_log(
            action='CREATE',
            table_name='sale',
            record_id=sale.id,
            new_values=sale.to_dict(),
            changes_summary=changes_summary
        )

        # Update daily stock record
        update_daily_stock_sales(variant.product_id, sale_date)

        # Update daily summary
        update_daily_summary(sale_date)

        db.session.commit()

        # Success messages
        success_msg = 'Sale recorded successfully!'
        if sale.discount_amount > 0:
            success_msg = f'Sale recorded with discount (KES {sale.discount_amount:.2f} off)!'

        flash(success_msg, 'success')

        if payment_total > sale.total_amount:
            change_amount = payment_total - sale.total_amount
            flash(f'Change given: KES {change_amount:,.2f}', 'info')

        # Redirect to the return_date if provided, otherwise use sale_date
        redirect_date = return_date if return_date else sale_date_str
        return redirect(url_for('sales', date=redirect_date))

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error adding sale: {str(e)}")
        flash(f'An error occurred while recording the sale: {str(e)}', 'error')
        redirect_date = request.form.get('return_date', date.today().strftime('%Y-%m-%d'))
        return redirect(url_for('sales', date=redirect_date))


@app.route('/edit_sale/<int:sale_id>', methods=['GET', 'POST'])
@login_required
def edit_sale(sale_id):
    sale = db.session.get(Sale, sale_id)
    if not sale:
        flash('Sale not found!', 'error')
        return redirect(url_for('sales'))

    current_user = get_current_user()

    if current_user.role not in ['admin', 'manager'] and sale.attendant_id != current_user.id:
        flash('You can only edit your own sales!', 'error')
        return redirect(url_for('sales'))

    if request.method == 'POST':
        try:
            old_values = sale.to_dict()
            old_quantity = sale.quantity
            original_sale_date = sale.sale_date

            # Update sale fields
            new_quantity = safe_float(request.form['quantity'])
            sale.unit_price = safe_float(request.form['unit_price'])
            sale.cash_amount = safe_float(request.form.get('cash_amount', 0))
            sale.mpesa_amount = safe_float(request.form.get('mpesa_amount', 0))
            sale.credit_amount = safe_float(request.form.get('credit_amount', 0))
            sale.customer_name = request.form.get('customer_name', '').strip() or None
            sale.notes = request.form.get('notes', '').strip() or None

            # Check if sale date changed
            new_sale_date_str = request.form.get('sale_date')
            if new_sale_date_str:
                try:
                    new_sale_date = datetime.strptime(new_sale_date_str, '%Y-%m-%d').date()
                    sale.sale_date = new_sale_date
                except ValueError:
                    flash('Invalid sale date format.', 'error')
                    variants = db.session.query(ProductVariant, Product, Category, Size).select_from(ProductVariant) \
                        .join(Product, ProductVariant.product_id == Product.id) \
                        .join(Category, Product.category_id == Category.id) \
                        .join(Size, ProductVariant.size_id == Size.id) \
                        .filter(ProductVariant.is_active == True) \
                        .order_by(Category.name, Product.name, Size.sort_order).all()
                    return render_template('sales/edit_sale.html', sale=sale, variants=variants)

            # Handle stock changes (convert to base units)
            quantity_diff = new_quantity - old_quantity
            if quantity_diff > 0:
                # Need more stock - convert to base units
                base_units_needed = quantity_diff * sale.variant.conversion_factor
                if not sale.variant.product.reduce_stock(base_units_needed):
                    flash(f'Insufficient stock to increase quantity by {quantity_diff}', 'error')
                    variants = db.session.query(ProductVariant, Product, Category, Size).select_from(ProductVariant) \
                        .join(Product, ProductVariant.product_id == Product.id) \
                        .join(Category, Product.category_id == Category.id) \
                        .join(Size, ProductVariant.size_id == Size.id) \
                        .filter(ProductVariant.is_active == True) \
                        .order_by(Category.name, Product.name, Size.sort_order).all()
                    return render_template('sales/edit_sale.html', sale=sale, variants=variants)
            elif quantity_diff < 0:
                # Return stock - convert to base units
                base_units_to_return = abs(quantity_diff) * sale.variant.conversion_factor
                sale.variant.product.add_stock(base_units_to_return)

            sale.quantity = new_quantity

            # Recalculate amounts
            sale.original_amount = sale.quantity * sale.unit_price
            sale.calculate_discount()

            # Update payment method
            payment_methods = []
            if sale.cash_amount > 0: payment_methods.append('cash')
            if sale.mpesa_amount > 0: payment_methods.append('mpesa')
            if sale.credit_amount > 0: payment_methods.append('credit')

            sale.payment_method = 'mixed' if len(payment_methods) > 1 else (
                payment_methods[0] if payment_methods else 'cash')

            new_values = sale.to_dict()
            changes_summary = get_changes_summary(old_values, new_values)
            create_audit_log(
                action='UPDATE',
                table_name='sale',
                record_id=sale.id,
                old_values=old_values,
                new_values=new_values,
                changes_summary=f"Sale updated: {sale.variant.get_display_name()} - {changes_summary}"
            )

            # Update daily stock for current sale date
            update_daily_stock_sales(sale.variant.product_id, sale.sale_date)

            # Update daily summaries for both dates if date changed
            if original_sale_date != sale.sale_date:
                update_daily_summary(original_sale_date)  # Update old date
                update_daily_summary(sale.sale_date)  # Update new date
            else:
                update_daily_summary(sale.sale_date)  # Update current date

            db.session.commit()
            flash('Sale updated successfully!', 'success')
            return redirect(url_for('sales', date=sale.sale_date.strftime('%Y-%m-%d')))

        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating sale: {str(e)}")
            flash(f'Error updating sale: {str(e)}', 'error')

    variants = db.session.query(ProductVariant, Product, Category, Size).select_from(ProductVariant) \
        .join(Product, ProductVariant.product_id == Product.id) \
        .join(Category, Product.category_id == Category.id) \
        .join(Size, ProductVariant.size_id == Size.id) \
        .filter(ProductVariant.is_active == True) \
        .order_by(Category.name, Product.name, Size.sort_order).all()
    return render_template('sales/edit_sale.html', sale=sale, variants=variants)


@app.route('/delete_sale/<int:sale_id>', methods=['POST'])
@login_required
def delete_sale(sale_id):
    try:
        sale = db.session.get(Sale, sale_id)
        if not sale:
            flash('Sale not found!', 'error')
            return redirect(url_for('sales'))

        current_user = get_current_user()

        if current_user.role not in ['admin', 'manager'] and sale.attendant_id != current_user.id:
            flash('You can only delete your own sales!', 'error')
            return redirect(url_for('sales'))

        old_values = sale.to_dict()
        variant_name = sale.variant.get_display_name() if sale.variant else 'Unknown'
        changes_summary = f"Sale deleted: {variant_name} x{sale.quantity} = KES {sale.total_amount}"
        if sale.customer_name:
            changes_summary += f" (Customer: {sale.customer_name})"

        # Return stock to product (convert to base units)
        base_units_to_return = sale.quantity * sale.variant.conversion_factor
        sale.variant.product.add_stock(base_units_to_return)

        product_id = sale.variant.product_id
        sale_date = sale.sale_date

        create_audit_log(
            action='DELETE',
            table_name='sale',
            record_id=sale_id,
            old_values=old_values,
            changes_summary=changes_summary
        )

        db.session.delete(sale)

        # Update daily stock
        update_daily_stock_sales(product_id, sale_date)

        # Update daily summary
        update_daily_summary(sale_date)

        db.session.commit()
        flash(f'Sale deleted successfully! {changes_summary}', 'success')

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error deleting sale: {str(e)}")
        flash(f'Error deleting sale: {str(e)}', 'error')

    return redirect(url_for('sales'))


# STOCK MANAGEMENT ROUTES - UPDATED VERSION
def get_or_create_daily_stock(product_id, stock_date):
    """Get or create daily stock record - FIXED VERSION"""
    daily_stock = DailyStock.query.filter_by(product_id=product_id, date=stock_date).first()

    if not daily_stock:
        product = db.session.get(Product, product_id)

        # Get yesterday's closing stock
        previous_date = stock_date - timedelta(days=1)
        previous_stock = DailyStock.query.filter_by(
            product_id=product_id,
            date=previous_date
        ).first()

        # CRITICAL FIX: Opening stock should ONLY come from previous day's closing
        # NEVER include current day's purchases in opening stock
        if previous_stock:
            opening_stock = previous_stock.closing_stock
        else:
            # If no previous record, use product's current stock as baseline
            opening_stock = product.current_stock if product else 0

        # Calculate additions from purchases ONLY (not included in opening)
        purchase_additions = db.session.query(
            db.func.coalesce(db.func.sum(StockPurchase.quantity), 0)
        ).filter(
            StockPurchase.product_id == product_id,
            StockPurchase.purchase_date == stock_date
        ).scalar() or 0

        # Calculate sales in base units for this date
        total_base_units_sold = db.session.query(
            db.func.coalesce(db.func.sum(Sale.quantity * ProductVariant.conversion_factor), 0)
        ).join(ProductVariant, Sale.variant_id == ProductVariant.id).filter(
            ProductVariant.product_id == product_id,
            Sale.sale_date == stock_date
        ).scalar() or 0

        daily_stock = DailyStock(
            product_id=product_id,
            date=stock_date,
            opening_stock=opening_stock,  # Only previous closing, NO purchases
            additions=purchase_additions,  # Purchases added separately
            sales_quantity=total_base_units_sold,
            closing_stock=0  # Will be calculated
        )

        daily_stock.calculate_closing_stock()
        db.session.add(daily_stock)
        db.session.flush()

    return daily_stock


@app.route('/add_stock_purchase', methods=['POST'])
@admin_required
def add_stock_purchase():
    """FIXED VERSION - Stock purchase should NEVER affect opening stock"""
    try:
        # Check if we should return to a specific date
        return_date = request.form.get('return_date')

        product_id = int(request.form.get('product_id'))
        quantity = safe_float(request.form.get('quantity'))
        unit_cost = safe_float(request.form.get('unit_cost'))
        supplier_name = request.form.get('supplier_name', '').strip()
        invoice_number = request.form.get('invoice_number', '').strip()
        purchase_date_str = request.form.get('purchase_date')
        notes = request.form.get('notes', '').strip()

        current_user = get_current_user()

        # Validate inputs
        if not product_id or quantity <= 0 or unit_cost <= 0:
            flash('Please fill in all required fields with valid values!', 'error')
            redirect_date = return_date if return_date else date.today().strftime('%Y-%m-%d')
            return redirect(url_for('daily_stock', date=redirect_date))

        # Parse date
        try:
            purchase_date = datetime.strptime(purchase_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            flash('Invalid purchase date format.', 'error')
            redirect_date = return_date if return_date else date.today().strftime('%Y-%m-%d')
            return redirect(url_for('daily_stock', date=redirect_date))

        # Get product
        product = db.session.get(Product, product_id)
        if not product:
            flash('Product not found!', 'error')
            redirect_date = return_date if return_date else date.today().strftime('%Y-%m-%d')
            return redirect(url_for('daily_stock', date=redirect_date))

        # Calculate total cost
        total_cost = quantity * unit_cost

        # Create purchase record
        purchase = StockPurchase(
            product_id=product_id,
            quantity=quantity,
            unit_cost=unit_cost,
            total_cost=total_cost,
            supplier_name=supplier_name if supplier_name else None,
            invoice_number=invoice_number if invoice_number else None,
            purchase_date=purchase_date,
            notes=notes if notes else None,
            recorded_by=current_user.id
        )

        db.session.add(purchase)
        db.session.flush()

        # CRITICAL FIX: Update product stock FIRST
        product.add_stock(quantity)

        # THEN get or create daily stock record
        # This ensures opening stock is correct (from previous day only)
        daily_stock = get_or_create_daily_stock(product_id, purchase_date)

        # Recalculate total additions from all purchases for this date
        total_purchase_additions = db.session.query(
            db.func.coalesce(db.func.sum(StockPurchase.quantity), 0)
        ).filter(
            StockPurchase.product_id == product_id,
            StockPurchase.purchase_date == purchase_date
        ).scalar() or 0

        # CRITICAL: Only update additions, NEVER touch opening_stock here
        daily_stock.additions = total_purchase_additions

        # Recalculate closing stock: opening + additions - sales
        daily_stock.calculate_closing_stock()

        daily_stock.updated_by = current_user.id
        daily_stock.updated_at = datetime.now(timezone.utc)

        # Create audit log
        changes_summary = f"Stock purchase: {product.name} x{quantity} @ KES {unit_cost} = KES {total_cost}"
        if supplier_name:
            changes_summary += f" from {supplier_name}"
        if invoice_number:
            changes_summary += f" (Invoice: {invoice_number})"

        create_audit_log(
            action='CREATE',
            table_name='stock_purchase',
            record_id=purchase.id,
            new_values=purchase.to_dict(),
            changes_summary=changes_summary
        )

        db.session.commit()

        flash(
            f'✅ Purchase recorded successfully! Added {quantity} {product.base_unit}s of {product.name} (KES {total_cost:,.2f})',
            'success')

        # Redirect to the return_date if provided, otherwise use purchase_date
        redirect_date = return_date if return_date else purchase_date_str
        return redirect(url_for('daily_stock', date=redirect_date))

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error recording stock purchase: {str(e)}")
        flash(f'Error recording stock purchase: {str(e)}', 'error')
        redirect_date = request.form.get('return_date', date.today().strftime('%Y-%m-%d'))
        return redirect(url_for('daily_stock', date=redirect_date))


@app.route('/daily_stock')
@login_required
def daily_stock():
    """FIXED VERSION - Better stock calculation logic"""
    selected_date_str = request.args.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
    except ValueError:
        selected_date = date.today()

    # Check if a specific date was requested (not today)
    stick_to_date = selected_date_str != date.today().strftime('%Y-%m-%d')

    # Get stock data
    stock_data = db.session.query(Product, Category, DailyStock).select_from(Product) \
        .join(Category, Product.category_id == Category.id) \
        .outerjoin(DailyStock, (Product.id == DailyStock.product_id) & (DailyStock.date == selected_date)) \
        .order_by(Category.name, Product.name).all()

    processed_stock_data = []
    for product, category, daily_stock in stock_data:
        if not daily_stock:
            # Create temporary stock record with correct logic
            temp_stock = DailyStock(product_id=product.id, date=selected_date)

            # Get previous day's closing stock (this is the ONLY source for opening)
            previous_date = selected_date - timedelta(days=1)
            previous_stock = DailyStock.query.filter_by(
                product_id=product.id,
                date=previous_date
            ).first()

            # CRITICAL FIX: Opening stock comes ONLY from previous day
            if previous_stock:
                temp_stock.opening_stock = previous_stock.closing_stock
            else:
                # If no previous record exists, use current product stock as baseline
                # This handles the initial setup or gaps in records
                temp_stock.opening_stock = product.current_stock

            # Calculate additions from purchases ONLY (separate from opening)
            purchase_additions = db.session.query(
                db.func.coalesce(db.func.sum(StockPurchase.quantity), 0)
            ).filter(
                StockPurchase.product_id == product.id,
                StockPurchase.purchase_date == selected_date
            ).scalar() or 0

            temp_stock.additions = purchase_additions

            # Calculate sales in base units for this date
            total_base_units_sold = db.session.query(
                db.func.coalesce(db.func.sum(Sale.quantity * ProductVariant.conversion_factor), 0)
            ).join(ProductVariant, Sale.variant_id == ProductVariant.id).filter(
                ProductVariant.product_id == product.id,
                Sale.sale_date == selected_date
            ).scalar() or 0

            temp_stock.sales_quantity = total_base_units_sold

            # Calculate closing: opening + additions - sales
            temp_stock.calculate_closing_stock()
            daily_stock = temp_stock
        else:
            # Verify and sync existing record
            purchase_additions = db.session.query(
                db.func.coalesce(db.func.sum(StockPurchase.quantity), 0)
            ).filter(
                StockPurchase.product_id == product.id,
                StockPurchase.purchase_date == selected_date
            ).scalar() or 0

            # CRITICAL: If additions don't match actual purchases, sync them
            if daily_stock.additions != purchase_additions:
                daily_stock.additions = purchase_additions
                daily_stock.calculate_closing_stock()

            # CRITICAL: Verify opening stock is correct (should match previous closing)
            previous_date = selected_date - timedelta(days=1)
            previous_stock = DailyStock.query.filter_by(
                product_id=product.id,
                date=previous_date
            ).first()

            if previous_stock and daily_stock.opening_stock != previous_stock.closing_stock:
                # Log the discrepancy
                app.logger.warning(
                    f"Opening stock mismatch for {product.name} on {selected_date}: "
                    f"Expected {previous_stock.closing_stock}, got {daily_stock.opening_stock}"
                )

        processed_stock_data.append((product, category, daily_stock))

    # Get purchases for this date
    daily_purchases = StockPurchase.query.filter_by(
        purchase_date=selected_date
    ).order_by(StockPurchase.timestamp.desc()).all()

    # Calculate total purchases for the day
    purchase_total = sum(purchase.total_cost for purchase in daily_purchases)

    # Get active products for the purchase form dropdown
    active_products = Product.query.join(Category).filter(
        Category.is_active == True
    ).order_by(Category.name, Product.name).all()

    return render_template('stock/daily_stock.html',
                           stock_data=processed_stock_data,
                           selected_date=selected_date,
                           stick_to_date=stick_to_date,
                           daily_purchases=daily_purchases,
                           purchase_total=purchase_total,
                           active_products=active_products,
                           today=date.today())


@app.route('/update_stock', methods=['POST'])
@admin_required
def update_stock():
    """Update daily stock opening balance - FIXED VERSION"""
    try:
        data = request.get_json()
        product_id = data.get('product_id')
        date_str = data.get('date')
        opening_stock = safe_float(data.get('opening_stock'))

        current_user = get_current_user()

        if not all([product_id, date_str]):
            return jsonify({'success': False, 'error': 'Missing required fields'}), 400

        # Parse date
        try:
            stock_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'error': 'Invalid date format'}), 400

        # Get or create daily stock
        daily_stock = get_or_create_daily_stock(product_id, stock_date)
        old_values = daily_stock.to_dict()

        # CRITICAL FIX: When manually adjusting opening stock,
        # we should log this as a correction, not a normal operation
        daily_stock.opening_stock = opening_stock
        daily_stock.updated_by = current_user.id
        daily_stock.updated_at = datetime.now(timezone.utc)

        # Recalculate closing stock based on new opening
        # Closing = Opening + Additions - Sales
        daily_stock.calculate_closing_stock()

        new_values = daily_stock.to_dict()
        changes_summary = get_changes_summary(old_values, new_values)

        # Create audit log
        create_audit_log(
            action='UPDATE',
            table_name='daily_stock',
            record_id=daily_stock.id,
            old_values=old_values,
            new_values=new_values,
            changes_summary=f"Manual opening stock adjustment for {daily_stock.product.name} on {stock_date} - {changes_summary}"
        )

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Opening stock updated successfully',
            'closing_stock': daily_stock.closing_stock,
            'warning': 'Opening stock manually adjusted. This should only be done to correct errors.'
        })

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error updating stock: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


# Add these routes to app.py

@app.route('/stock_overview')
@admin_required
def stock_overview():
    """Comprehensive stock overview and valuation"""

    # Get all products with their stock
    products = Product.query.join(Category).filter(Category.is_active == True).all()

    # Initialize statistics
    stock_stats = {
        'total_stock_value': 0,
        'potential_revenue': 0,
        'potential_profit': 0,
        'profit_margin': 0,
        'total_units': 0,
        'total_products': len(products),
        'active_products': len([p for p in products if p.current_stock > 0]),
        'low_stock_count': 0,
        'low_stock_value': 0,
        'out_of_stock_count': 0,
        'good_stock_count': 0,
        'good_stock_value': 0
    }

    # Calculate stock by category
    stock_by_category = []
    category_data = {}

    for product in products:
        stock = product.get_available_stock()
        stock_value = stock * product.base_buying_price

        # Calculate potential revenue from variants
        variants = product.get_active_variants()
        if variants:
            avg_selling_price = sum(v.selling_price * v.conversion_factor for v in variants) / len(variants)
            potential_revenue = stock * avg_selling_price
        else:
            potential_revenue = 0

        potential_profit = potential_revenue - stock_value

        # Update totals
        stock_stats['total_stock_value'] += stock_value
        stock_stats['potential_revenue'] += potential_revenue
        stock_stats['potential_profit'] += potential_profit
        stock_stats['total_units'] += stock

        # Stock status
        if stock <= 0:
            stock_stats['out_of_stock_count'] += 1
        elif stock <= product.min_stock_level:
            stock_stats['low_stock_count'] += 1
            stock_stats['low_stock_value'] += stock_value
        else:
            stock_stats['good_stock_count'] += 1
            stock_stats['good_stock_value'] += stock_value

        # Group by category
        cat_name = product.category.name
        if cat_name not in category_data:
            category_data[cat_name] = {
                'name': cat_name,
                'product_count': 0,
                'total_units': 0,
                'stock_value': 0,
                'potential_revenue': 0,
                'potential_profit': 0
            }

        category_data[cat_name]['product_count'] += 1
        category_data[cat_name]['total_units'] += stock
        category_data[cat_name]['stock_value'] += stock_value
        category_data[cat_name]['potential_revenue'] += potential_revenue
        category_data[cat_name]['potential_profit'] += potential_profit

    # Calculate profit margin
    if stock_stats['potential_revenue'] > 0:
        stock_stats['profit_margin'] = (stock_stats['potential_profit'] / stock_stats['potential_revenue']) * 100

    # Convert category data to sorted list
    stock_by_category = sorted(category_data.values(), key=lambda x: x['stock_value'], reverse=True)

    # Top 10 most valuable products
    top_products = []
    for product in products:
        stock = product.get_available_stock()
        if stock > 0:
            stock_value = stock * product.base_buying_price
            variants = product.get_active_variants()
            avg_selling_price = sum(v.selling_price * v.conversion_factor for v in variants) / len(
                variants) if variants else 0
            potential_revenue = stock * avg_selling_price
            potential_profit = potential_revenue - stock_value

            top_products.append({
                'id': product.id,
                'name': product.name,
                'category': product.category.name,
                'stock': stock,
                'base_unit': product.base_unit,
                'cost_per_unit': product.base_buying_price,
                'stock_value': stock_value,
                'avg_selling_price': avg_selling_price,
                'potential_revenue': potential_revenue,
                'potential_profit': potential_profit,
                'min_stock': product.min_stock_level
            })

    top_products = sorted(top_products, key=lambda x: x['stock_value'], reverse=True)[:10]

    # Get low stock and out of stock products
    low_stock_products = Product.query.filter(
        Product.current_stock > 0,
        Product.current_stock <= Product.min_stock_level
    ).order_by(Product.current_stock.asc()).all()

    out_of_stock_products = Product.query.filter(
        Product.current_stock <= 0
    ).order_by(Product.name).all()

    # Prepare chart data
    category_chart_data = {
        'labels': [cat['name'] for cat in stock_by_category],
        'values': [cat['stock_value'] for cat in stock_by_category]
    }

    return render_template('stock/stock_overview.html',
                           stock_stats=stock_stats,
                           stock_by_category=stock_by_category,
                           top_products=top_products,
                           low_stock_products=low_stock_products,
                           out_of_stock_products=out_of_stock_products,
                           category_chart_data=category_chart_data,
                           now=datetime.now())


# Update the stock_purchases route to include overall stats
@app.route('/stock_purchases')
@admin_required
def stock_purchases():
    """View all stock purchases with statistics"""
    selected_date_str = request.args.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
    except ValueError:
        selected_date = date.today()

    # Get purchases for the selected date
    purchases = StockPurchase.query.filter_by(
        purchase_date=selected_date
    ).order_by(StockPurchase.timestamp.desc()).all()

    # Calculate DAILY purchase statistics
    purchase_stats = {
        'total_value': 0,
        'total_units': 0,
        'avg_unit_cost': 0,
        'unique_products': 0,
        'unique_suppliers': 0,
        'highest_purchase': 0,
        'potential_revenue': 0,
        'top_products_by_value': [],
        'top_products_by_quantity': []
    }

    if purchases:
        purchase_stats['total_value'] = sum(p.total_cost for p in purchases)
        purchase_stats['total_units'] = sum(p.quantity for p in purchases)
        purchase_stats['avg_unit_cost'] = (purchase_stats['total_value'] / purchase_stats['total_units']) if \
        purchase_stats['total_units'] > 0 else 0
        purchase_stats['unique_products'] = len(set(p.product_id for p in purchases))
        purchase_stats['unique_suppliers'] = len(set(p.supplier_name for p in purchases if p.supplier_name))
        purchase_stats['highest_purchase'] = max(p.total_cost for p in purchases)

        # Calculate potential revenue
        potential_revenue = 0
        for purchase in purchases:
            variants = purchase.product.get_active_variants()
            if variants:
                avg_selling_price = sum(v.selling_price * v.conversion_factor for v in variants) / len(variants)
                potential_revenue += purchase.quantity * avg_selling_price
        purchase_stats['potential_revenue'] = potential_revenue

        # Group by product
        from collections import defaultdict
        product_data = defaultdict(lambda: {'total_cost': 0, 'total_quantity': 0, 'name': ''})

        for purchase in purchases:
            product_id = purchase.product_id
            product_data[product_id]['total_cost'] += purchase.total_cost
            product_data[product_id]['total_quantity'] += purchase.quantity
            product_data[product_id]['name'] = purchase.product.name

        product_list = [
            {
                'name': data['name'],
                'total_cost': data['total_cost'],
                'total_quantity': data['total_quantity']
            }
            for data in product_data.values()
        ]

        purchase_stats['top_products_by_value'] = sorted(product_list, key=lambda x: x['total_cost'], reverse=True)
        purchase_stats['top_products_by_quantity'] = sorted(product_list, key=lambda x: x['total_quantity'],
                                                            reverse=True)

    # Calculate OVERALL stock statistics (for the summary card)
    products = Product.query.join(Category).filter(Category.is_active == True).all()

    overall_stock_stats = {
        'total_stock_value': 0,
        'potential_revenue': 0,
        'potential_profit': 0
    }

    for product in products:
        stock = product.get_available_stock()
        stock_value = stock * product.base_buying_price

        variants = product.get_active_variants()
        if variants:
            avg_selling_price = sum(v.selling_price * v.conversion_factor for v in variants) / len(variants)
            potential_revenue = stock * avg_selling_price
        else:
            potential_revenue = 0

        overall_stock_stats['total_stock_value'] += stock_value
        overall_stock_stats['potential_revenue'] += potential_revenue

    overall_stock_stats['potential_profit'] = overall_stock_stats['potential_revenue'] - overall_stock_stats[
        'total_stock_value']

    # Get active products for dropdown
    products = Product.query.order_by(Product.name).all()

    return render_template('stock/stock_purchases.html',
                           purchases=purchases,
                           selected_date=selected_date,
                           products=products,
                           today=date.today(),
                           purchase_stats=purchase_stats,
                           overall_stock_stats=overall_stock_stats)

@app.route('/edit_stock_purchase/<int:purchase_id>', methods=['GET', 'POST'])
@admin_required
def edit_stock_purchase(purchase_id):
    """Edit a stock purchase"""
    purchase = db.session.get(StockPurchase, purchase_id)
    if not purchase:
        flash('Stock purchase not found!', 'error')
        return redirect(url_for('stock_purchases'))

    if request.method == 'POST':
        try:
            old_values = purchase.to_dict()
            old_quantity = purchase.quantity
            old_date = purchase.purchase_date

            # Update purchase fields
            new_quantity = safe_float(request.form.get('quantity'))
            purchase.unit_cost = safe_float(request.form.get('unit_cost'))
            purchase.supplier_name = request.form.get('supplier_name', '').strip() or None
            purchase.invoice_number = request.form.get('invoice_number', '').strip() or None
            purchase.notes = request.form.get('notes', '').strip() or None

            # Parse new date
            new_date_str = request.form.get('purchase_date')
            try:
                new_date = datetime.strptime(new_date_str, '%Y-%m-%d').date()
                purchase.purchase_date = new_date
            except (ValueError, TypeError):
                flash('Invalid purchase date format.', 'error')
                products = Product.query.order_by(Product.name).all()
                return render_template('stock/edit_stock_purchase.html', purchase=purchase, products=products)

            # Recalculate total cost
            purchase.total_cost = new_quantity * purchase.unit_cost

            # Adjust product stock
            quantity_diff = new_quantity - old_quantity
            if quantity_diff != 0:
                if quantity_diff > 0:
                    purchase.product.add_stock(quantity_diff)
                else:
                    purchase.product.reduce_stock(abs(quantity_diff))

            purchase.quantity = new_quantity

            # Update daily stock for old date
            if old_date == new_date:
                daily_stock = get_or_create_daily_stock(purchase.product_id, old_date)
                # Recalculate additions for this date
                total_additions = db.session.query(
                    db.func.coalesce(db.func.sum(StockPurchase.quantity), 0)
                ).filter(
                    StockPurchase.product_id == purchase.product_id,
                    StockPurchase.purchase_date == old_date
                ).scalar()
                daily_stock.additions = total_additions
                daily_stock.calculate_closing_stock()
            else:
                # Update both dates if date changed
                for target_date in [old_date, new_date]:
                    daily_stock = get_or_create_daily_stock(purchase.product_id, target_date)
                    total_additions = db.session.query(
                        db.func.coalesce(db.func.sum(StockPurchase.quantity), 0)
                    ).filter(
                        StockPurchase.product_id == purchase.product_id,
                        StockPurchase.purchase_date == target_date
                    ).scalar()
                    daily_stock.additions = total_additions
                    daily_stock.calculate_closing_stock()

            new_values = purchase.to_dict()
            changes_summary = get_changes_summary(old_values, new_values)

            create_audit_log(
                action='UPDATE',
                table_name='stock_purchase',
                record_id=purchase.id,
                old_values=old_values,
                new_values=new_values,
                changes_summary=f"Stock purchase updated: {purchase.product.name} - {changes_summary}"
            )

            db.session.commit()
            flash('Stock purchase updated successfully!', 'success')
            return redirect(url_for('stock_purchases', date=purchase.purchase_date.strftime('%Y-%m-%d')))

        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating stock purchase: {str(e)}")
            flash(f'Error updating stock purchase: {str(e)}', 'error')

    products = Product.query.order_by(Product.name).all()
    return render_template('stock/edit_stock_purchase.html', purchase=purchase, products=products)


@app.route('/delete_stock_purchase/<int:purchase_id>', methods=['POST'])
@admin_required
def delete_stock_purchase(purchase_id):
    """Delete a stock purchase"""
    try:
        purchase = db.session.get(StockPurchase, purchase_id)
        if not purchase:
            flash('Stock purchase not found!', 'error')
            return redirect(url_for('stock_purchases'))

        old_values = purchase.to_dict()
        product_name = purchase.product.name if purchase.product else 'Unknown'
        purchase_date = purchase.purchase_date
        product_id = purchase.product_id
        quantity = purchase.quantity

        changes_summary = f"Stock purchase deleted: {product_name} x{quantity} = KES {purchase.total_cost}"

        # Reduce product stock
        purchase.product.reduce_stock(quantity)

        # Create audit log before deletion
        create_audit_log(
            action='DELETE',
            table_name='stock_purchase',
            record_id=purchase_id,
            old_values=old_values,
            changes_summary=changes_summary
        )

        # Delete purchase
        db.session.delete(purchase)

        # Update daily stock
        daily_stock = get_or_create_daily_stock(product_id, purchase_date)
        total_additions = db.session.query(
            db.func.coalesce(db.func.sum(StockPurchase.quantity), 0)
        ).filter(
            StockPurchase.product_id == product_id,
            StockPurchase.purchase_date == purchase_date
        ).scalar()
        daily_stock.additions = total_additions
        daily_stock.calculate_closing_stock()

        db.session.commit()
        flash(f'Stock purchase deleted successfully! {changes_summary}', 'success')

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error deleting stock purchase: {str(e)}")
        flash(f'Error deleting stock purchase: {str(e)}', 'error')

    return redirect(url_for('stock_purchases'))



        
# EXPENSES ROUTES
@app.route('/expenses')
@login_required
def expenses():
    selected_date_str = request.args.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
    except ValueError:
        selected_date = date.today()

    current_user = get_current_user()

    query = db.session.query(Expense, User, ExpenseCategory).select_from(Expense) \
        .join(User, Expense.recorded_by == User.id) \
        .join(ExpenseCategory, Expense.expense_category_id == ExpenseCategory.id) \
        .filter(Expense.expense_date == selected_date)

    if current_user.role not in ['admin', 'manager']:
        query = query.filter(Expense.recorded_by == current_user.id)

    expenses_data = query.order_by(Expense.timestamp.desc()).all()
    total_expenses = sum(expense.amount for expense, _, _ in expenses_data)

    expense_categories = ExpenseCategory.query.filter_by(is_active=True).order_by(ExpenseCategory.name).all()

    return render_template('expenses/expenses.html',
                           expenses_data=expenses_data,
                           selected_date=selected_date,
                           total_expenses=total_expenses,
                           expense_categories=expense_categories,
                           current_user=current_user)


@app.route('/add_expense', methods=['POST'])
@login_required
def add_expense():
    try:
        # Check if we should return to a specific date
        return_date = request.form.get('return_date')

        # Extract form data
        description = request.form['description'].strip()
        amount = safe_float(request.form['amount'])
        expense_category_id = int(request.form['expense_category_id'])
        expense_date_str = request.form['expense_date']
        notes = request.form.get('notes', '').strip()

        current_user = get_current_user()

        # Validate required fields
        if not all([description, expense_category_id]) or amount <= 0:
            flash('Please fill in all required fields with valid values!', 'error')
            redirect_date = return_date if return_date else expense_date_str
            return redirect(url_for('expenses', date=redirect_date))

        # Parse expense date
        try:
            expense_date = datetime.strptime(expense_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            flash('Invalid expense date format!', 'error')
            redirect_date = return_date if return_date else date.today().strftime('%Y-%m-%d')
            return redirect(url_for('expenses', date=redirect_date))

        # Validate expense category
        expense_category = db.session.get(ExpenseCategory, expense_category_id)
        if not expense_category or not expense_category.is_active:
            flash('Invalid expense category selected!', 'error')
            redirect_date = return_date if return_date else expense_date_str
            return redirect(url_for('expenses', date=redirect_date))

        # Create expense record
        expense = Expense(
            description=description,
            amount=amount,
            expense_category_id=expense_category_id,
            expense_date=expense_date,
            notes=notes if notes else None,
            recorded_by=current_user.id
        )

        db.session.add(expense)
        db.session.flush()

        # Create audit log with detailed summary
        changes_summary = f"Expense recorded: {description} - KES {amount:,.2f} ({expense_category.name})"
        if notes:
            changes_summary += f" - Notes: {notes}"

        create_audit_log(
            action='CREATE',
            table_name='expense',
            record_id=expense.id,
            new_values=expense.to_dict(),
            changes_summary=changes_summary
        )

        # Update daily summary for the expense date
        update_daily_summary(expense_date)

        db.session.commit()

        flash(f'✅ Expense recorded successfully! {changes_summary}', 'success')

        # Redirect to the return_date if provided (sticky date), otherwise use expense_date
        redirect_date = return_date if return_date else expense_date_str
        return redirect(url_for('expenses', date=redirect_date))

    except ValueError as ve:
        db.session.rollback()
        flash('Invalid expense date format!', 'error')
        app.logger.error(f"ValueError in add_expense: {str(ve)}")
        redirect_date = request.form.get('return_date', date.today().strftime('%Y-%m-%d'))
        return redirect(url_for('expenses', date=redirect_date))

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error recording expense: {str(e)}")
        flash(f'Error recording expense: {str(e)}', 'error')
        redirect_date = request.form.get('return_date', date.today().strftime('%Y-%m-%d'))
        return redirect(url_for('expenses', date=redirect_date))

@app.route('/edit_expense/<int:expense_id>', methods=['GET', 'POST'])
@login_required
def edit_expense(expense_id):
    expense = db.session.get(Expense, expense_id)
    if not expense:
        flash('Expense not found!', 'error')
        return redirect(url_for('expenses'))

    current_user = get_current_user()

    if current_user.role not in ['admin', 'manager'] and expense.recorded_by != current_user.id:
        flash('You can only edit your own expenses!', 'error')
        return redirect(url_for('expenses'))

    if request.method == 'POST':
        try:
            old_values = expense.to_dict()
            original_expense_date = expense.expense_date

            expense.description = request.form['description'].strip()
            expense.amount = safe_float(request.form['amount'])
            expense.expense_category_id = int(request.form['expense_category_id'])
            expense.expense_date = datetime.strptime(request.form['expense_date'], '%Y-%m-%d').date()
            expense.notes = request.form.get('notes', '').strip() or None

            if not all([expense.description, expense.expense_category_id]) or expense.amount <= 0:
                flash('Please fill in all required fields with valid values!', 'error')
                expense_categories = ExpenseCategory.query.filter_by(is_active=True).order_by(
                    ExpenseCategory.name).all()
                return render_template('expenses/edit_expense.html', expense=expense, expense_categories=expense_categories)

            # Validate expense category
            expense_category = db.session.get(ExpenseCategory, expense.expense_category_id)
            if not expense_category or not expense_category.is_active:
                flash('Invalid expense category selected!', 'error')
                expense_categories = ExpenseCategory.query.filter_by(is_active=True).order_by(
                    ExpenseCategory.name).all()
                return render_template('expenses/edit_expense.html', expense=expense, expense_categories=expense_categories)

            new_values = expense.to_dict()

            changes_summary = get_changes_summary(old_values, new_values)
            create_audit_log(
                action='UPDATE',
                table_name='expense',
                record_id=expense.id,
                old_values=old_values,
                new_values=new_values,
                changes_summary=f"Expense updated: {expense.description} - {changes_summary}"
            )

            # Update daily summaries for both dates if date changed
            if original_expense_date != expense.expense_date:
                update_daily_summary(original_expense_date)  # Update old date
                update_daily_summary(expense.expense_date)  # Update new date
            else:
                update_daily_summary(expense.expense_date)  # Update current date

            db.session.commit()
            flash('Expense updated successfully!', 'success')
            return redirect(url_for('expenses'))

        except ValueError:
            flash('Invalid expense date format!', 'error')
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating expense: {str(e)}")
            flash(f'Error updating expense: {str(e)}', 'error')

    expense_categories = ExpenseCategory.query.filter_by(is_active=True).order_by(ExpenseCategory.name).all()
    return render_template('expenses/edit_expense.html', expense=expense, expense_categories=expense_categories)


@app.route('/delete_expense/<int:expense_id>', methods=['POST'])
@login_required
def delete_expense(expense_id):
    try:
        expense = db.session.get(Expense, expense_id)
        if not expense:
            flash('Expense not found!', 'error')
            return redirect(url_for('expenses'))

        current_user = get_current_user()

        if current_user.role not in ['admin', 'manager'] and expense.recorded_by != current_user.id:
            flash('You can only delete your own expenses!', 'error')
            return redirect(url_for('expenses'))

        old_values = expense.to_dict()
        category_name = expense.expense_category.name if expense.expense_category else 'Unknown'
        changes_summary = f"Expense deleted: {expense.description} - KES {expense.amount} ({category_name})"
        expense_date = expense.expense_date

        create_audit_log(
            action='DELETE',
            table_name='expense',
            record_id=expense_id,
            old_values=old_values,
            changes_summary=changes_summary
        )

        db.session.delete(expense)

        # Update daily summary
        update_daily_summary(expense_date)

        db.session.commit()
        flash(f'Expense deleted successfully! {changes_summary}', 'success')

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error deleting expense: {str(e)}")
        flash(f'Error deleting expense: {str(e)}', 'error')

    return redirect(url_for('expenses'))


# REPORTS ROUTES
# Add these imports at the top of app.py
import csv
from io import StringIO, BytesIO
from flask import make_response
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# Add these routes to app.py (replace or add to existing reports route)

@app.route('/reports')
@login_required
def reports():
    start_date_str = request.args.get('start_date', (date.today().replace(day=1)).strftime('%Y-%m-%d'))
    end_date_str = request.args.get('end_date', date.today().strftime('%Y-%m-%d'))

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Invalid date format!', 'error')
        start_date = date.today().replace(day=1)
        end_date = date.today()

    current_user = get_current_user()

    # Daily summaries
    daily_summaries = DailySummary.query.filter(
        DailySummary.date.between(start_date, end_date)
    ).order_by(DailySummary.date.desc()).all()

    # Product sales with updated variant system
    product_query = db.session.query(
        Product.name,
        Category.name.label('category_name'),
        db.func.coalesce(db.func.sum(Sale.quantity), 0).label('total_quantity'),
        db.func.coalesce(db.func.sum(Sale.total_amount), 0).label('total_sales'),
        db.func.coalesce(db.func.sum(
            Sale.total_amount - (Product.base_buying_price * ProductVariant.conversion_factor * Sale.quantity)),
                         0).label('total_profit')
    ).select_from(Product) \
        .join(ProductVariant, Product.id == ProductVariant.product_id) \
        .join(Sale, ProductVariant.id == Sale.variant_id) \
        .join(Category, Product.category_id == Category.id) \
        .filter(Sale.sale_date.between(start_date, end_date))

    if current_user.role not in ['admin', 'manager']:
        product_query = product_query.filter(Sale.attendant_id == current_user.id)

    product_sales = product_query.group_by(Product.id, Product.name, Category.name) \
        .order_by(db.text('total_sales DESC')).all()

    # Category-wise sales for pie chart
    category_sales = db.session.query(
        Category.name,
        db.func.coalesce(db.func.sum(Sale.total_amount), 0).label('total_sales')
    ).select_from(Category) \
        .join(Product, Category.id == Product.category_id) \
        .join(ProductVariant, Product.id == ProductVariant.product_id) \
        .join(Sale, ProductVariant.id == Sale.variant_id) \
        .filter(Sale.sale_date.between(start_date, end_date))

    if current_user.role not in ['admin', 'manager']:
        category_sales = category_sales.filter(Sale.attendant_id == current_user.id)

    category_sales = category_sales.group_by(Category.name).all()

    # Payment method breakdown
    payment_breakdown = db.session.query(
        db.func.coalesce(db.func.sum(Sale.cash_amount), 0).label('cash'),
        db.func.coalesce(db.func.sum(Sale.mpesa_amount), 0).label('mpesa'),
        db.func.coalesce(db.func.sum(Sale.credit_amount), 0).label('credit')
    ).filter(Sale.sale_date.between(start_date, end_date))

    if current_user.role not in ['admin', 'manager']:
        payment_breakdown = payment_breakdown.filter(Sale.attendant_id == current_user.id)

    payment_breakdown = payment_breakdown.first()

    # Expense breakdown by category
    expense_breakdown = db.session.query(
        ExpenseCategory.name,
        db.func.coalesce(db.func.sum(Expense.amount), 0).label('total_amount')
    ).select_from(ExpenseCategory) \
        .join(Expense, ExpenseCategory.id == Expense.expense_category_id) \
        .filter(Expense.expense_date.between(start_date, end_date))

    if current_user.role not in ['admin', 'manager']:
        expense_breakdown = expense_breakdown.filter(Expense.recorded_by == current_user.id)

    expense_breakdown = expense_breakdown.group_by(ExpenseCategory.name).all()

    # Attendant performance
    attendant_performance = []
    if current_user.role in ['admin', 'manager']:
        attendant_performance = db.session.query(
            User.full_name,
            db.func.count(Sale.id).label('total_transactions'),
            db.func.coalesce(db.func.sum(Sale.quantity), 0).label('total_quantity'),
            db.func.coalesce(db.func.sum(Sale.total_amount), 0).label('total_sales'),
            db.func.coalesce(db.func.sum(
                Sale.total_amount - (Product.base_buying_price * ProductVariant.conversion_factor * Sale.quantity)),
                             0).label('total_profit')
        ).select_from(User) \
            .join(Sale, User.id == Sale.attendant_id) \
            .join(ProductVariant, Sale.variant_id == ProductVariant.id) \
            .join(Product, ProductVariant.product_id == Product.id) \
            .filter(Sale.sale_date.between(start_date, end_date)) \
            .group_by(User.id, User.full_name) \
            .order_by(db.text('total_sales DESC')).all()

    # Calculate period totals
    if current_user.role in ['admin', 'manager']:
        total_period_sales = sum(summary.total_sales or 0 for summary in daily_summaries)
        total_period_profit = sum(summary.total_profit or 0 for summary in daily_summaries)
        total_period_expenses = sum(summary.total_expenses or 0 for summary in daily_summaries)
    else:
        personal_metrics = db.session.query(
            db.func.coalesce(db.func.sum(Sale.total_amount), 0).label('sales'),
            db.func.coalesce(db.func.sum(
                Sale.total_amount - (Product.base_buying_price * ProductVariant.conversion_factor * Sale.quantity)),
                             0).label('profit')
        ).join(ProductVariant).join(Product).filter(
            Sale.attendant_id == current_user.id,
            Sale.sale_date.between(start_date, end_date)
        ).first()

        personal_expenses = db.session.query(
            db.func.coalesce(db.func.sum(Expense.amount), 0)
        ).filter(
            Expense.recorded_by == current_user.id,
            Expense.expense_date.between(start_date, end_date)
        ).scalar()

        total_period_sales = personal_metrics.sales if personal_metrics else 0
        total_period_profit = personal_metrics.profit if personal_metrics else 0
        total_period_expenses = personal_expenses or 0

    total_period_net_profit = total_period_profit - total_period_expenses

    return render_template('reports/reports.html',
                           daily_summaries=daily_summaries,
                           product_sales=product_sales,
                           category_sales=category_sales,
                           payment_breakdown=payment_breakdown,
                           expense_breakdown=expense_breakdown,
                           attendant_performance=attendant_performance,
                           start_date=start_date,
                           end_date=end_date,
                           total_period_sales=total_period_sales,
                           total_period_profit=total_period_profit,
                           total_period_expenses=total_period_expenses,
                           total_period_net_profit=total_period_net_profit,
                           current_user=current_user)


@app.route('/reports/export/<export_type>')
@login_required
def export_report(export_type):
    """Export reports to CSV or Excel"""
    start_date_str = request.args.get('start_date', (date.today().replace(day=1)).strftime('%Y-%m-%d'))
    end_date_str = request.args.get('end_date', date.today().strftime('%Y-%m-%d'))
    format_type = request.args.get('format', 'csv')  # csv or excel

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Invalid date format!', 'error')
        return redirect(url_for('reports'))

    current_user = get_current_user()

    if export_type == 'sales':
        return export_sales_report(start_date, end_date, format_type, current_user)
    elif export_type == 'products':
        return export_products_report(start_date, end_date, format_type, current_user)
    elif export_type == 'expenses':
        return export_expenses_report(start_date, end_date, format_type, current_user)
    elif export_type == 'daily_summary':
        return export_daily_summary_report(start_date, end_date, format_type, current_user)
    elif export_type == 'full':
        return export_full_report(start_date, end_date, format_type, current_user)
    else:
        flash('Invalid export type!', 'error')
        return redirect(url_for('reports'))


def export_sales_report(start_date, end_date, format_type, current_user):
    """Export sales data"""
    query = db.session.query(
        Sale.sale_date,
        Product.name.label('product_name'),
        Size.name.label('size_name'),
        Sale.quantity,
        Sale.unit_price,
        Sale.original_amount,
        Sale.discount_amount,
        Sale.total_amount,
        Sale.payment_method,
        User.full_name.label('attendant_name'),
        Sale.customer_name
    ).select_from(Sale) \
        .join(ProductVariant, Sale.variant_id == ProductVariant.id) \
        .join(Product, ProductVariant.product_id == Product.id) \
        .join(Size, ProductVariant.size_id == Size.id) \
        .join(User, Sale.attendant_id == User.id) \
        .filter(Sale.sale_date.between(start_date, end_date))

    if current_user.role not in ['admin', 'manager']:
        query = query.filter(Sale.attendant_id == current_user.id)

    sales_data = query.order_by(Sale.sale_date.desc()).all()

    if format_type == 'excel':
        return create_excel_report('Sales Report', [
            'Date', 'Product', 'Size', 'Quantity', 'Unit Price', 'Original Amount',
            'Discount', 'Total Amount', 'Payment Method', 'Attendant', 'Customer'
        ], sales_data, f'sales_report_{start_date}_{end_date}.xlsx')
    else:
        return create_csv_report('Sales Report', [
            'Date', 'Product', 'Size', 'Quantity', 'Unit Price', 'Original Amount',
            'Discount', 'Total Amount', 'Payment Method', 'Attendant', 'Customer'
        ], sales_data, f'sales_report_{start_date}_{end_date}.csv')


def export_products_report(start_date, end_date, format_type, current_user):
    """Export product sales summary"""
    query = db.session.query(
        Product.name.label('product_name'),
        Category.name.label('category_name'),
        db.func.coalesce(db.func.sum(Sale.quantity), 0).label('total_quantity'),
        db.func.coalesce(db.func.sum(Sale.total_amount), 0).label('total_sales'),
        db.func.coalesce(db.func.sum(
            Sale.total_amount - (Product.base_buying_price * ProductVariant.conversion_factor * Sale.quantity)),
                         0).label('total_profit')
    ).select_from(Product) \
        .join(ProductVariant, Product.id == ProductVariant.product_id) \
        .join(Sale, ProductVariant.id == Sale.variant_id) \
        .join(Category, Product.category_id == Category.id) \
        .filter(Sale.sale_date.between(start_date, end_date))

    if current_user.role not in ['admin', 'manager']:
        query = query.filter(Sale.attendant_id == current_user.id)

    product_data = query.group_by(Product.id, Product.name, Category.name) \
        .order_by(db.text('total_sales DESC')).all()

    headers = ['Product', 'Category', 'Quantity Sold', 'Total Sales', 'Total Profit'] if current_user.role in ['admin',
                                                                                                               'manager'] else [
        'Product', 'Category', 'Quantity Sold', 'Total Sales']

    if format_type == 'excel':
        return create_excel_report('Product Sales Report', headers, product_data,
                                   f'product_sales_{start_date}_{end_date}.xlsx')
    else:
        return create_csv_report('Product Sales Report', headers, product_data,
                                 f'product_sales_{start_date}_{end_date}.csv')


def export_expenses_report(start_date, end_date, format_type, current_user):
    """Export expenses data"""
    query = db.session.query(
        Expense.expense_date,
        ExpenseCategory.name.label('category_name'),
        Expense.description,
        Expense.amount,
        User.full_name.label('recorded_by_name'),
        Expense.notes
    ).select_from(Expense) \
        .join(ExpenseCategory, Expense.expense_category_id == ExpenseCategory.id) \
        .join(User, Expense.recorded_by == User.id) \
        .filter(Expense.expense_date.between(start_date, end_date))

    if current_user.role not in ['admin', 'manager']:
        query = query.filter(Expense.recorded_by == current_user.id)

    expense_data = query.order_by(Expense.expense_date.desc()).all()

    if format_type == 'excel':
        return create_excel_report('Expenses Report', [
            'Date', 'Category', 'Description', 'Amount', 'Recorded By', 'Notes'
        ], expense_data, f'expenses_report_{start_date}_{end_date}.xlsx')
    else:
        return create_csv_report('Expenses Report', [
            'Date', 'Category', 'Description', 'Amount', 'Recorded By', 'Notes'
        ], expense_data, f'expenses_report_{start_date}_{end_date}.csv')


def export_daily_summary_report(start_date, end_date, format_type, current_user):
    """Export daily summary"""
    daily_summaries = DailySummary.query.filter(
        DailySummary.date.between(start_date, end_date)
    ).order_by(DailySummary.date.desc()).all()

    data = [(s.date, s.total_transactions, s.total_sales, s.total_profit, s.total_expenses, s.net_profit) for s in
            daily_summaries]

    if format_type == 'excel':
        return create_excel_report('Daily Summary Report', [
            'Date', 'Transactions', 'Total Sales', 'Gross Profit', 'Expenses', 'Net Profit'
        ], data, f'daily_summary_{start_date}_{end_date}.xlsx')
    else:
        return create_csv_report('Daily Summary Report', [
            'Date', 'Transactions', 'Total Sales', 'Gross Profit', 'Expenses', 'Net Profit'
        ], data, f'daily_summary_{start_date}_{end_date}.csv')


def export_full_report(start_date, end_date, format_type, current_user):
    """Export comprehensive Excel report with multiple sheets"""
    if format_type != 'excel':
        flash('Full report is only available in Excel format', 'warning')
        return redirect(url_for('reports'))

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Add Summary Sheet
    ws_summary = wb.create_sheet('Summary')
    add_summary_sheet(ws_summary, start_date, end_date, current_user)

    # Add Sales Sheet
    ws_sales = wb.create_sheet('Sales')
    add_sales_sheet(ws_sales, start_date, end_date, current_user)

    # Add Products Sheet
    ws_products = wb.create_sheet('Product Sales')
    add_products_sheet(ws_products, start_date, end_date, current_user)

    # Add Expenses Sheet
    ws_expenses = wb.create_sheet('Expenses')
    add_expenses_sheet(ws_expenses, start_date, end_date, current_user)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.read())
    response.headers['Content-Disposition'] = f'attachment; filename=full_report_{start_date}_{end_date}.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


def create_csv_report(title, headers, data, filename):
    """Create CSV report"""
    si = StringIO()
    writer = csv.writer(si)

    writer.writerow([title])
    writer.writerow([])
    writer.writerow(headers)

    for row in data:
        writer.writerow(row)

    output = make_response(si.getvalue())
    output.headers['Content-Disposition'] = f'attachment; filename={filename}'
    output.headers['Content-Type'] = 'text/csv'

    return output


def create_excel_report(title, headers, data, filename):
    """Create Excel report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    # Title
    ws.merge_cells('A1:' + chr(65 + len(headers) - 1) + '1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')

    # Headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data
    for row_idx, row_data in enumerate(data, 4):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.read())
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


def add_summary_sheet(ws, start_date, end_date, current_user):
    """Add summary information to worksheet"""
    ws.append(['Business Summary Report'])
    ws.append([f'Period: {start_date} to {end_date}'])
    ws.append([])

    # Get totals
    if current_user.role in ['admin', 'manager']:
        daily_summaries = DailySummary.query.filter(
            DailySummary.date.between(start_date, end_date)
        ).all()
        total_sales = sum(s.total_sales or 0 for s in daily_summaries)
        total_profit = sum(s.total_profit or 0 for s in daily_summaries)
        total_expenses = sum(s.total_expenses or 0 for s in daily_summaries)
    else:
        # Personal metrics for attendants
        total_sales = 0
        total_profit = 0
        total_expenses = 0

    ws.append(['Metric', 'Amount'])
    ws.append(['Total Sales', total_sales])
    ws.append(['Gross Profit', total_profit])
    ws.append(['Total Expenses', total_expenses])
    ws.append(['Net Profit', total_profit - total_expenses])


def add_sales_sheet(ws, start_date, end_date, current_user):
    """Add sales data to worksheet"""
    query = db.session.query(
        Sale.sale_date,
        Product.name,
        Size.name,
        Sale.quantity,
        Sale.unit_price,
        Sale.total_amount,
        User.full_name
    ).select_from(Sale) \
        .join(ProductVariant, Sale.variant_id == ProductVariant.id) \
        .join(Product, ProductVariant.product_id == Product.id) \
        .join(Size, ProductVariant.size_id == Size.id) \
        .join(User, Sale.attendant_id == User.id) \
        .filter(Sale.sale_date.between(start_date, end_date))

    if current_user.role not in ['admin', 'manager']:
        query = query.filter(Sale.attendant_id == current_user.id)

    sales_data = query.order_by(Sale.sale_date.desc()).all()

    ws.append(['Date', 'Product', 'Size', 'Quantity', 'Unit Price', 'Total', 'Attendant'])
    for sale in sales_data:
        ws.append(list(sale))


def add_products_sheet(ws, start_date, end_date, current_user):
    """Add product sales data to worksheet"""
    query = db.session.query(
        Product.name,
        Category.name,
        db.func.sum(Sale.quantity),
        db.func.sum(Sale.total_amount)
    ).select_from(Product) \
        .join(ProductVariant, Product.id == ProductVariant.product_id) \
        .join(Sale, ProductVariant.id == Sale.variant_id) \
        .join(Category, Product.category_id == Category.id) \
        .filter(Sale.sale_date.between(start_date, end_date))

    if current_user.role not in ['admin', 'manager']:
        query = query.filter(Sale.attendant_id == current_user.id)

    product_data = query.group_by(Product.id, Product.name, Category.name).all()

    ws.append(['Product', 'Category', 'Quantity Sold', 'Total Sales'])
    for product in product_data:
        ws.append(list(product))


def add_expenses_sheet(ws, start_date, end_date, current_user):
    """Add expenses data to worksheet"""
    query = db.session.query(
        Expense.expense_date,
        ExpenseCategory.name,
        Expense.description,
        Expense.amount
    ).select_from(Expense) \
        .join(ExpenseCategory, Expense.expense_category_id == ExpenseCategory.id) \
        .filter(Expense.expense_date.between(start_date, end_date))

    if current_user.role not in ['admin', 'manager']:
        query = query.filter(Expense.recorded_by == current_user.id)

    expense_data = query.order_by(Expense.expense_date.desc()).all()

    ws.append(['Date', 'Category', 'Description', 'Amount'])
    for expense in expense_data:
        ws.append(list(expense))

# USER PROFILE ROUTES
@app.route('/profile')
@login_required
def profile():
    current_user = get_current_user()
    today = date.today()
    this_month = today.replace(day=1)

    user_stats = {
        'today_sales': db.session.query(
            db.func.count(Sale.id),
            db.func.coalesce(db.func.sum(Sale.total_amount), 0)
        ).filter(
            Sale.attendant_id == current_user.id,
            Sale.sale_date == today
        ).first(),
        'month_sales': db.session.query(
            db.func.count(Sale.id),
            db.func.coalesce(db.func.sum(Sale.total_amount), 0)
        ).filter(
            Sale.attendant_id == current_user.id,
            Sale.sale_date >= this_month
        ).first(),
        'total_sales': db.session.query(
            db.func.count(Sale.id),
            db.func.coalesce(db.func.sum(Sale.total_amount), 0)
        ).filter(
            Sale.attendant_id == current_user.id
        ).first()
    }

    return render_template('profile.html', user=current_user, user_stats=user_stats)


@app.route('/change_password', methods=['POST'])
@login_required
def change_password():
    current_user = get_current_user()
    current_password = request.form['current_password']
    new_password = request.form['new_password']
    confirm_password = request.form['confirm_password']

    if not current_user.check_password(current_password):
        flash('Current password is incorrect', 'error')
        return redirect(url_for('profile'))

    if new_password != confirm_password:
        flash('New passwords do not match', 'error')
        return redirect(url_for('profile'))

    if len(new_password) < 6:
        flash('Password must be at least 6 characters long', 'error')
        return redirect(url_for('profile'))

    create_audit_log(
        action='UPDATE',
        table_name='user',
        record_id=current_user.id,
        changes_summary=f"Password changed by {current_user.full_name}"
    )

    current_user.set_password(new_password)
    db.session.commit()

    flash('Password changed successfully!', 'success')
    return redirect(url_for('profile'))


# AUDIT LOGS ROUTES
@app.route('/audit_logs')
@admin_required
def audit_logs():
    page = request.args.get('page', 1, type=int)
    per_page = min(request.args.get('per_page', 50, type=int), 100)

    filters = {
        'action': request.args.get('action', 'all'),
        'table': request.args.get('table', 'all'),
        'user': request.args.get('user', 'all'),
        'start_date': request.args.get('start_date', ''),
        'end_date': request.args.get('end_date', '')
    }

    query = db.session.query(AuditLog, User).join(User, AuditLog.user_id == User.id)

    if filters['action'] != 'all':
        query = query.filter(AuditLog.action == filters['action'].upper())

    if filters['table'] != 'all':
        query = query.filter(AuditLog.table_name == filters['table'])

    if filters['user'] != 'all':
        query = query.filter(AuditLog.user_id == int(filters['user']))

    if filters['start_date']:
        try:
            start_datetime = datetime.strptime(filters['start_date'], '%Y-%m-%d')
            query = query.filter(AuditLog.timestamp >= start_datetime)
        except ValueError:
            pass

    if filters['end_date']:
        try:
            end_datetime = datetime.strptime(filters['end_date'], '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            query = query.filter(AuditLog.timestamp <= end_datetime)
        except ValueError:
            pass

    query = query.order_by(AuditLog.timestamp.desc())
    audit_logs = query.paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )

    actions = db.session.query(AuditLog.action).distinct().all()
    tables = db.session.query(AuditLog.table_name).distinct().all()
    users = User.query.filter_by(is_active=True).order_by(User.full_name).all()

    return render_template('audit/audit_logs.html',
                           audit_logs=audit_logs,
                           actions=[a[0] for a in actions],
                           tables=[t[0] for t in tables],
                           users=users,
                           current_filters=filters)


@app.route('/audit_log/<int:log_id>')
@admin_required
def audit_log_detail(log_id):
    """View detailed information about a specific audit log entry"""
    audit_log = db.session.get(AuditLog, log_id)
    if not audit_log:
        flash('Audit log entry not found!', 'error')
        return redirect(url_for('audit_logs'))

    # Parse JSON values if they exist
    old_values = None
    new_values = None

    if audit_log.old_values:
        try:
            old_values = json.loads(audit_log.old_values)
        except (json.JSONDecodeError, TypeError):
            old_values = audit_log.old_values

    if audit_log.new_values:
        try:
            new_values = json.loads(audit_log.new_values)
        except (json.JSONDecodeError, TypeError):
            new_values = audit_log.new_values

    return render_template('audit/audit_log_detail.html',
                           audit_log=audit_log,
                           old_values=old_values,
                           new_values=new_values)


# API ROUTES
@app.route('/api/products')
@login_required
def api_products():
    products = db.session.query(Product, Category).select_from(Product) \
        .join(Category, Product.category_id == Category.id) \
        .filter(Category.is_active == True).all()

    return jsonify([{
        'id': p.id,
        'name': p.name,
        'category_name': category.name,
        'base_unit': p.base_unit,
        'base_buying_price': p.base_buying_price,
        'current_stock': p.get_available_stock(),
        'stock_status': p.get_stock_status(),
        'min_stock_level': p.min_stock_level
    } for p, category in products])


@app.route('/api/product_variants/<int:product_id>')
@login_required
def api_product_variants(product_id):
    variants = db.session.query(ProductVariant, Size).join(Size).filter(
        ProductVariant.product_id == product_id,
        ProductVariant.is_active == True,
        Size.is_active == True
    ).order_by(Size.sort_order).all()

    return jsonify([{
        'id': v.id,
        'display_name': v.get_display_name(),
        'size_name': size.name,
        'selling_price': v.selling_price,
        'conversion_factor': v.conversion_factor,
        'available_quantity': v.get_available_stock_in_variant_units(),
        'profit_per_unit': v.get_profit_per_unit()
    } for v, size in variants])


@app.route('/api/variant_stock/<int:variant_id>')
@login_required
def api_variant_stock(variant_id):
    variant = db.session.get(ProductVariant, variant_id)
    if not variant:
        return jsonify({'error': 'Variant not found'}), 404

    return jsonify({
        'available_quantity': variant.get_available_stock_in_variant_units(),
        'base_stock': variant.product.get_available_stock(),
        'display_name': variant.get_display_name(),
        'selling_price': variant.selling_price,
        'profit_per_unit': variant.get_profit_per_unit()
    })


@app.route('/api/discount_permissions')
@login_required
def api_discount_permissions():
    current_user = get_current_user()
    max_percentage_map = {'attendant': 10, 'manager': 25, 'admin': 100}

    return jsonify({
        'can_give_discount': True,
        'max_percentage': max_percentage_map.get(current_user.role, 0),
        'requires_reason': True
    })


#Search functions
@app.route('/search/suggestions')
@login_required
def search_suggestions():
    """API endpoint for search suggestions"""
    query = request.args.get('q', '').strip()
    limit = min(int(request.args.get('limit', 5)), 20)

    if len(query) < 2:
        return jsonify({'suggestions': []})

    suggestions = []
    current_user = get_current_user()

    # Search products
    products = Product.query.join(Category).filter(
        Product.name.ilike(f'%{query}%'),
        Category.is_active == True
    ).limit(limit).all()

    for product in products:
        suggestions.append({
            'type': 'product',
            'text': product.name,
            'category': product.category.name,
            'icon': 'fas fa-box',
            'url': url_for('product_variants', product_id=product.id)  # Direct link to variants
        })

    # Search product variants by size
    variants = db.session.query(ProductVariant, Product, Size).select_from(ProductVariant) \
        .join(Product, ProductVariant.product_id == Product.id) \
        .join(Size, ProductVariant.size_id == Size.id) \
        .join(Category, Product.category_id == Category.id) \
        .filter(
        (Product.name.ilike(f'%{query}%')) | (Size.name.ilike(f'%{query}%')),
        ProductVariant.is_active == True,
        Category.is_active == True
    ).limit(limit).all()

    for variant, product, size in variants:
        suggestions.append({
            'type': 'variant',
            'text': f"{product.name} - {size.name}",
            'category': f"KES {variant.selling_price:,.2f}",
            'icon': 'fas fa-wine-bottle',
            'url': url_for('product_variants', product_id=product.id)
        })

    # Search categories (admin/manager only)
    if current_user.role in ['admin', 'manager']:
        categories = Category.query.filter(
            Category.name.ilike(f'%{query}%'),
            Category.is_active == True
        ).limit(3).all()

        for category in categories:
            suggestions.append({
                'type': 'category',
                'text': category.name,
                'category': 'Product Category',
                'icon': 'fas fa-layer-group',
                'url': url_for('categories')
            })

    # Search users (admin only)
    if current_user.role == 'admin':
        users = User.query.filter(
            User.full_name.ilike(f'%{query}%') | User.username.ilike(f'%{query}%'),
            User.is_active == True
        ).limit(3).all()

        for user in users:
            suggestions.append({
                'type': 'user',
                'text': user.full_name,
                'category': f'{user.role.title()} - {user.username}',
                'icon': 'fas fa-user',
                'url': url_for('users')
            })

    return jsonify({'suggestions': suggestions[:limit]})


@app.route('/search')
@login_required
def search():
    """Full search results page"""
    query = request.args.get('q', '').strip()
    search_type = request.args.get('type', 'all')

    if not query:
        flash('Please enter a search term', 'warning')
        return redirect(url_for('dashboard'))

    current_user = get_current_user()
    results = {
        'products': [],
        'variants': [],
        'categories': [],
        'users': [],
        'sales': []
    }

    # Search products
    if search_type in ['all', 'product']:
        results['products'] = Product.query.join(Category).filter(
            Product.name.ilike(f'%{query}%'),
            Category.is_active == True
        ).all()

    # Search product variants
    if search_type in ['all', 'variant']:
        results['variants'] = db.session.query(ProductVariant, Product, Size).select_from(ProductVariant) \
            .join(Product, ProductVariant.product_id == Product.id) \
            .join(Size, ProductVariant.size_id == Size.id) \
            .join(Category, Product.category_id == Category.id) \
            .filter(
            (Product.name.ilike(f'%{query}%')) | (Size.name.ilike(f'%{query}%')),
            ProductVariant.is_active == True,
            Category.is_active == True
        ).all()

    # Search categories (admin/manager only)
    if current_user.role in ['admin', 'manager'] and search_type in ['all', 'category']:
        results['categories'] = Category.query.filter(
            Category.name.ilike(f'%{query}%')
        ).all()

    # Search users (admin only)
    if current_user.role == 'admin' and search_type in ['all', 'user']:
        results['users'] = User.query.filter(
            User.full_name.ilike(f'%{query}%') | User.username.ilike(f'%{query}%')
        ).all()

    # Search recent sales
    if search_type in ['all', 'sale']:
        sales_query = db.session.query(Sale, ProductVariant, Product).select_from(Sale) \
            .join(ProductVariant, Sale.variant_id == ProductVariant.id) \
            .join(Product, ProductVariant.product_id == Product.id) \
            .filter(Product.name.ilike(f'%{query}%'))

        if current_user.role not in ['admin', 'manager']:
            sales_query = sales_query.filter(Sale.attendant_id == current_user.id)

        results['sales'] = sales_query.order_by(Sale.timestamp.desc()).limit(20).all()

    return render_template('search_results.html',
                           query=query,
                           results=results,
                           search_type=search_type)


# ERROR HANDLERS
@app.errorhandler(404)
def not_found(error):
    current_user = get_current_user()
    if current_user:
        create_audit_log(
            action='ERROR',
            table_name='system',
            changes_summary=f"404 Error - Page not found: {request.url}"
        )
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
    return render_template('errors/404.html'), 404


@app.errorhandler(403)
def forbidden(error):
    current_user = get_current_user()
    if current_user:
        create_audit_log(
            action='ERROR',
            table_name='system',
            changes_summary=f"403 Error - Access forbidden: {request.url}"
        )
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
    return render_template('errors/403.html'), 403


@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    current_user = get_current_user()
    if current_user:
        create_audit_log(
            action='ERROR',
            table_name='system',
            changes_summary=f"500 Error - Internal server error: {request.url} - {str(error)}"
        )
        try:
            db.session.commit()
        except Exception:
            pass
    return render_template('errors/500.html'), 500


# TEMPLATE GLOBALS
@app.template_global()
def format_currency(amount):
    if amount is None:
        return "KES 0"
    try:
        return "KES {:,.0f}".format(float(amount))
    except (ValueError, TypeError):
        return "KES 0"


@app.template_global()
def get_stock_status(stock_level, min_level=5):
    try:
        stock_level = float(stock_level or 0)
        min_level = float(min_level or 5)

        if stock_level <= 0:
            return "out_of_stock"
        elif stock_level <= min_level:
            return "low_stock"
        elif stock_level <= min_level * 2:
            return "medium_stock"
        else:
            return "good_stock"
    except (ValueError, TypeError):
        return "unknown"


@app.template_global()
def get_stock_status_color(stock_level, min_level=5):
    status = get_stock_status(stock_level, min_level)
    color_map = {
        'out_of_stock': 'danger',
        'low_stock': 'warning',
        'medium_stock': 'info',
        'good_stock': 'success',
        'unknown': 'secondary'
    }
    return color_map.get(status, 'secondary')


@app.template_global()
def get_stock_icon(stock_level, min_level=5):
    status = get_stock_status(stock_level, min_level)
    icon_map = {
        'out_of_stock': 'exclamation-triangle',
        'low_stock': 'exclamation-circle',
        'medium_stock': 'info-circle',
        'good_stock': 'check-circle',
        'unknown': 'question-circle'
    }
    return icon_map.get(status, 'question-circle')


# Add utility functions to Jinja environment
app.jinja_env.globals.update({
    'min': min,
    'max': max,
    'abs': abs,
    'round': round,
    'len': len,
    'safe_float': safe_float
})


# DATABASE INITIALIZATION
def initialize_database():
    try:
        with app.app_context():
            db.create_all()

            if User.query.count() == 0:
                # Create default users
                default_users = [
                    {
                        'username': 'admin',
                        'email': 'admin@liquorstore.com',
                        'full_name': 'System Administrator',
                        'role': 'admin',
                        'password': 'admin123'
                    },
                    {
                        'username': 'manager1',
                        'email': 'manager@liquorstore.com',
                        'full_name': 'Jane Smith',
                        'role': 'manager',
                        'password': 'manager123'
                    },
                    {
                        'username': 'attendant1',
                        'email': 'attendant1@liquorstore.com',
                        'full_name': 'John Doe',
                        'role': 'attendant',
                        'password': 'attendant123'
                    },
                    {
                        'username': 'attendant2',
                        'email': 'attendant2@liquorstore.com',
                        'full_name': 'Mary Johnson',
                        'role': 'attendant',
                        'password': 'attendant123'
                    }
                ]

                for user_data in default_users:
                    user = User(
                        username=user_data['username'],
                        email=user_data['email'],
                        full_name=user_data['full_name'],
                        role=user_data['role']
                    )
                    user.set_password(user_data['password'])
                    db.session.add(user)

                db.session.commit()

                print("Database initialized with default users!")
                print("\nDefault login credentials:")
                for user_data in default_users:
                    print(
                        f"{user_data['role'].title()}: username='{user_data['username']}', password='{user_data['password']}'"
                    )
                print("\n*** REMEMBER TO CHANGE DEFAULT PASSWORDS IN PRODUCTION! ***")

            else:
                print("Database already contains users.")

    except Exception as e:
        print(f"Error initializing database: {str(e)}")
        raise


@app.context_processor
def inject_user():
    return dict(current_user=get_current_user())


def create_app(config=None):
    if config:
        app.config.update(config)

    with app.app_context():
        initialize_database()

    return app




if __name__ == '__main__':
    initialize_database()
    app.run(debug=True, host='0.0.0.0', port=5000)